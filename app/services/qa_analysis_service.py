"""AI analysis for QA logs stored as SQLite BLOBs."""
from __future__ import annotations

import io
import json
import re
import sqlite3
import time
from collections.abc import Callable
from pathlib import Path
from typing import Any

import release_system.core as core
from app.db.connection import transaction
from app.repositories import qa_repo
from app.repositories.audit_repo import log_audit
from app.timeutil import beijing_timestamp

_QA_TEST_RESULT_STATUSES = {"pass", "fail", "skip", "unknown"}

_QA_ANALYSIS_SYSTEM = (
    "You analyze QA test logs/summary tables for an HPC release. Match the "
    "log content to the provided per-app test inventory and report results.\n"
    "Return STRICT JSON of the form: "
    "{\"apps\": [{\"app_id\": str, "
    "\"qa_status\": \"qa_passed\"|\"has_issues\"|\"cannot_release\"|\"not_checked\", "
    "\"qa_issue_note\": str, \"tests\": [{\"test\": str, \"arch\": str, "
    "\"status\": \"pass\"|\"fail\"|\"skip\"|\"unknown\", \"perf\": str, \"note\": str}]}]}.\n"
    "Rules:\n"
    "- Include EVERY app_id from the inventory. If the log has no data for an app, set its tests to status=unknown and qa_status=not_checked.\n"
    "- qa_status=qa_passed only when every listed test for that app shows a clear pass in the log.\n"
    "- qa_status=has_issues when any test fails or shows clear regression; qa_issue_note must concisely list which tests/arches failed and why (in Chinese).\n"
    "- qa_status=cannot_release only when the log explicitly says release is blocked.\n"
    "- Match app/test names case-insensitively and tolerate small spelling variants. The log may be plain text OR a multi-sheet spreadsheet rendered as TSV (each sheet preceded by `### Sheet: <name> ###`).\n"
    "- perf: short numeric/throughput summary if the log contains one, else empty.\n"
    "- note: at most one short Chinese sentence per test (cause of fail, perf delta, or empty).\n"
    "- Do NOT invent tests not in the inventory. Output JSON only, no prose, no code fences."
)


def _analysis_inventory(
    release: dict[str, Any],
    apps: dict[str, dict[str, Any]],
) -> list[dict[str, Any]]:
    inventory: list[dict[str, Any]] = []
    for app_id, snapshot in release["snapshots"].items():
        if snapshot.get("release_decision") != "release":
            continue
        app = apps.get(app_id) or {}
        tests = []
        for test in snapshot.get("tests") or []:
            if not test.get("enabled", True):
                continue
            tests.append(
                {
                    "name": test.get("name") or test.get("path") or "",
                    "path": test.get("path") or "",
                    "command": test.get("command") or "",
                    "supported_chips": test.get("supported_chips") or [],
                    "arches": test.get("arch_list") or [],
                }
            )
        inventory.append(
            {
                "app_id": app_id,
                "app_name": snapshot.get("app_name") or app.get("name") or app_id,
                "version": snapshot.get("version") or snapshot.get("app_version") or "",
                "tests": tests,
            }
        )
    return inventory


def _xlsx_to_text(raw: bytes, *, max_rows_per_sheet: int = 2000) -> str:
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
    chunks: list[str] = []
    try:
        for name in workbook.sheetnames:
            sheet = workbook[name]
            chunks.append(f"### Sheet: {name} ###")
            for index, row in enumerate(sheet.iter_rows(values_only=True)):
                if index >= max_rows_per_sheet:
                    chunks.append(f"...[truncated after {max_rows_per_sheet} rows]...")
                    break
                cells = [
                    "" if value is None else str(value).replace("\t", " ").replace("\r", " ").replace("\n", " ")
                    for value in row
                ]
                while cells and cells[-1] == "":
                    cells.pop()
                if cells:
                    chunks.append("\t".join(cells))
            chunks.append("")
    finally:
        workbook.close()
    return "\n".join(chunks)


def _log_to_text(filename: str, raw: bytes) -> str:
    if Path(filename).suffix.lower() == ".xlsx":
        try:
            return _xlsx_to_text(raw)
        except Exception as exc:
            raise RuntimeError(f"无法解析 xlsx：{exc}") from exc
    return raw.decode("utf-8", errors="replace")


def _progress(
    callback: Callable[..., None] | None,
    stage: str,
    message: str,
    **extra: Any,
) -> None:
    if callback is None:
        return
    if extra:
        try:
            callback(stage, message, **extra)
            return
        except TypeError:
            pass
    callback(stage, message)


def _normalize_test_results(raw: Any) -> list[dict[str, str]]:
    if not isinstance(raw, list):
        return []
    cleaned: list[dict[str, str]] = []
    for row in raw:
        if not isinstance(row, dict):
            continue
        status = str(row.get("status") or "unknown").lower()
        if status not in _QA_TEST_RESULT_STATUSES:
            status = "unknown"
        cleaned.append(
            {
                "test": str(row.get("test") or "").strip(),
                "arch": str(row.get("arch") or "").strip(),
                "status": status,
                "perf": str(row.get("perf") or "").strip(),
                "note": str(row.get("note") or "").strip(),
            }
        )
    return cleaned


def analyze_qa_log(
    conn: sqlite3.Connection,
    release_id: str,
    *,
    llm_call: Callable[[str, str], str] | None = None,
    max_log_chars: int = 200_000,
    progress: Callable[..., None] | None = None,
    max_llm_attempts: int = 2,
) -> dict[str, Any]:
    """Analyze the uploaded log without materializing it on the filesystem."""
    _progress(progress, "checking_log", "正在检查已上传的 QA log")
    metadata = qa_repo.get_qa_log(conn, release_id)
    if not metadata:
        raise RuntimeError("本 release 还未上传 QA log")
    stored = qa_repo.get_qa_log_content(conn, release_id)
    if stored is None:
        raise RuntimeError("QA log 内容未存入数据库，请重新上传")
    raw, filename = stored

    _progress(progress, "reading_log", f"正在读取 QA log：{filename}")
    if Path(filename).suffix.lower() == ".xlsx":
        _progress(progress, "parsing_excel", "正在解析 Excel 文件")
    else:
        _progress(progress, "parsing_text", "正在解析文本 log")
    text = _log_to_text(filename, raw)
    truncated = False
    if len(text) > max_log_chars:
        _progress(progress, "truncating_log", "log 较大，正在截断中段以控制 LLM 上下文")
        half = max_log_chars // 2
        text = text[:half] + "\n\n...[log truncated]...\n\n" + text[-half:]
        truncated = True

    _progress(progress, "building_inventory", "正在准备 app 和测试清单")
    release = core.get_release(conn, release_id)
    apps = {app["id"]: app for app in core.list_apps(conn)}
    inventory = _analysis_inventory(release, apps)
    if not inventory:
        raise RuntimeError("本 release 没有 release 决策为 release 的 app，无法分析")

    _progress(progress, "building_prompt", "正在构造 LLM 分析上下文")
    user_payload = json.dumps({"inventory": inventory, "log": text}, ensure_ascii=False)
    if llm_call is None:
        from app.integrations.llm import chat_json

        def default_llm_call(system: str, payload: str) -> str:
            def stream_progress(token_count: int) -> None:
                _progress(
                    progress,
                    "streaming_llm",
                    f"正在接收 LLM 输出：已收到 {token_count} token",
                    token_count=token_count,
                )

            return chat_json(system, payload, progress=stream_progress)

        llm_call = default_llm_call

    parsed: dict[str, Any] | None = None
    max_llm_attempts = max(1, max_llm_attempts)
    for attempt in range(1, max_llm_attempts + 1):
        try:
            suffix = f"（第 {attempt}/{max_llm_attempts} 次）" if max_llm_attempts > 1 else ""
            _progress(progress, "waiting_llm", f"正在等待 LLM 返回结果{suffix}", token_count=0)
            reply = llm_call(_QA_ANALYSIS_SYSTEM, user_payload)
            _progress(progress, "parsing_llm", "正在解析 LLM 返回结果")
            try:
                parsed = json.loads(reply)
            except json.JSONDecodeError:
                cleaned = re.sub(
                    r"^```(?:json)?\s*|\s*```$",
                    "",
                    reply.strip(),
                    flags=re.MULTILINE,
                )
                parsed = json.loads(cleaned)
            break
        except Exception as exc:
            if attempt >= max_llm_attempts:
                raise
            _progress(progress, "retrying_llm", f"LLM 调用失败，正在重试：{exc}")
            time.sleep(min(attempt, 3))

    if parsed is None:
        raise RuntimeError("LLM 未返回有效结果")

    _progress(progress, "normalizing_result", "正在整理 AI 建议")
    valid_ids = {entry["app_id"] for entry in inventory}
    apps_out: list[dict[str, Any]] = []
    for row in parsed.get("apps") or []:
        if not isinstance(row, dict) or row.get("app_id") not in valid_ids:
            continue
        status = row.get("qa_status") or "not_checked"
        if status not in core.QA_STATUSES:
            status = "not_checked"
        apps_out.append(
            {
                "app_id": row["app_id"],
                "qa_status": status,
                "qa_issue_note": str(row.get("qa_issue_note") or "").strip(),
                "test_results": _normalize_test_results(row.get("tests")),
            }
        )

    ts = beijing_timestamp()
    with transaction(conn):
        log_audit(
            conn,
            f"QA AI 分析 log：{filename}",
            ts=ts,
            user="qa-ai",
            role="QA",
            release_id=release_id,
            event="qa_analyze_log",
        )
    _progress(progress, "completed", "AI 分析完成")
    return {"apps": apps_out, "log_truncated": truncated, "log_chars": len(raw)}
