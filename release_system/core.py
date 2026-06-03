from __future__ import annotations

import csv
import datetime as dt
import hashlib
import io
import json
import os
import re
import secrets
import sqlite3
import threading
import time
import uuid
from contextlib import contextmanager
from pathlib import Path
from typing import Any, Callable, Iterator


_INIT_LOCK = threading.Lock()
_INITIALIZED_DBS: set[str] = set()


class ManagedConnection(sqlite3.Connection):
    """SQLite connection that can defer helper-level commits in core.transaction."""

    def __init__(self, *args: Any, **kwargs: Any) -> None:
        super().__init__(*args, **kwargs)
        self._transaction_depth = 0

    def commit(self) -> None:
        if self._transaction_depth:
            return
        super().commit()

    def commit_now(self) -> None:
        super().commit()


@contextmanager
def transaction(conn: sqlite3.Connection) -> Iterator[None]:
    """Commit once on success and rollback all pending writes on failure.

    Connections returned by core.connect suppress nested helper commit() calls
    inside this context, so legacy helpers that still call commit() cannot split
    the transaction.
    """
    if not isinstance(conn, ManagedConnection):
        try:
            yield
        except Exception:
            conn.rollback()
            raise
        else:
            conn.commit()
        return

    if conn._transaction_depth:
        savepoint = f"core_tx_{conn._transaction_depth}_{uuid.uuid4().hex}"
        conn.execute(f"SAVEPOINT {savepoint}")
        conn._transaction_depth += 1
        try:
            yield
        except Exception:
            conn.execute(f"ROLLBACK TO SAVEPOINT {savepoint}")
            conn.execute(f"RELEASE SAVEPOINT {savepoint}")
            raise
        else:
            conn.execute(f"RELEASE SAVEPOINT {savepoint}")
        finally:
            conn._transaction_depth -= 1
        return

    conn._transaction_depth = 1
    try:
        if not conn.in_transaction:
            conn.execute("BEGIN")
        yield
    except Exception:
        conn.rollback()
        raise
    else:
        conn.commit_now()
    finally:
        conn._transaction_depth = 0


BEIJING_TZ = dt.timezone(dt.timedelta(hours=8))


def now() -> str:
    return dt.datetime.now(dt.timezone.utc).replace(microsecond=0).isoformat()


def beijing_now() -> dt.datetime:
    """Current Beijing time as a naive datetime (no tzinfo)."""
    return dt.datetime.now(BEIJING_TZ).replace(tzinfo=None, microsecond=0)


def normalize_deadline(value: str | None) -> str:
    """Normalize a deadline string to ``YYYY-MM-DD HH:MM`` (Beijing time).

    Accepts ``''`` (returns ``''``), ``YYYY-MM-DD``, ``YYYY-MM-DDTHH:MM[:SS]``,
    or ``YYYY-MM-DD HH:MM[:SS]``. Empty deadline means "no deadline set".
    """
    text = (value or "").strip()
    if not text:
        return ""
    text = text.replace("T", " ")
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            parsed = dt.datetime.strptime(text, fmt)
            if fmt == "%Y-%m-%d":
                parsed = parsed.replace(hour=23, minute=59)
            return parsed.strftime("%Y-%m-%d %H:%M")
        except ValueError:
            continue
    raise ValueError(f"Invalid deadline: {value!r}; expected YYYY-MM-DD or YYYY-MM-DD HH:MM")


def parse_deadline(value: str | None) -> dt.datetime | None:
    if not value:
        return None
    return dt.datetime.strptime(normalize_deadline(value), "%Y-%m-%d %H:%M")


def validate_deadline_order(app_freeze_deadline: str | None, doc_deadline: str | None) -> None:
    """Reject deadlines where app freeze lands after the doc deadline.

    Empty deadlines mean "not set" and are always accepted. current_phase
    assumes app freeze precedes the doc deadline; a reversed pair produces
    incoherent phases, so it is blocked at every entry point that sets them.
    """
    freeze = parse_deadline(app_freeze_deadline)
    doc = parse_deadline(doc_deadline)
    if freeze is not None and doc is not None and freeze > doc:
        raise ValueError(
            f"App 冻结 deadline（{normalize_deadline(app_freeze_deadline)}）"
            f"不能晚于 Doc deadline（{normalize_deadline(doc_deadline)}）"
        )


def is_before(deadline: str | None, *, ref: dt.datetime | None = None) -> bool:
    """True if the reference moment is strictly before the deadline.

    Empty/None deadline means "no deadline set" → treated as infinite future,
    so this returns True (i.e. the action is still allowed).
    """
    dl = parse_deadline(deadline)
    if dl is None:
        return True
    return (ref or beijing_now()) < dl


def current_phase(release: dict[str, Any]) -> str:
    """Derive the lifecycle phase of a release from its deadlines and lock flag."""
    if release.get("released_locked"):
        return "released_locked"
    if not is_before(release.get("doc_deadline", "")):
        return "after_doc_deadline"
    if not is_before(release.get("app_freeze_deadline", "")):
        return "after_app_freeze"
    return "before_app_freeze"


PHASES = ("before_app_freeze", "after_app_freeze", "after_doc_deadline", "released_locked")

# Single source of truth for "what is allowed in each release phase". Entry
# points (core helpers + server handlers) consult this table instead of
# re-deriving rules from is_before(...) checks; that way every action's phase
# gating stays consistent and changes only need to land here.
_PHASE_POLICY: dict[str, set[str]] = {
    "before_app_freeze": {
        "new_app_release", "new_app_non_release",
        "raise_to_release", "lower_decision",
        "edit_app_info", "expand_qa_scope",
        "edit_snapshot", "qa_set_status", "qa_upload_log",
    },
    "after_app_freeze": {
        "new_app_non_release",
        "lower_decision",
        "edit_app_info",
        "edit_snapshot", "qa_set_status", "qa_upload_log",
    },
    "after_doc_deadline": {
        "new_app_non_release",
        "lower_decision",
        "qa_set_status", "qa_upload_log",
    },
    "released_locked": set(),
}


def can(release_or_phase: dict[str, Any] | str, action: str) -> bool:
    """True if *action* is allowed in the given release's current phase.

    Accepts either a release dict (phase is derived) or a phase string. Unknown
    actions return False so a typo at a call site fails closed rather than
    silently allowing the write.
    """
    if isinstance(release_or_phase, dict):
        phase = current_phase(release_or_phase)
    else:
        phase = str(release_or_phase)
    return action in _PHASE_POLICY.get(phase, set())


def require_can(release: dict[str, Any], action: str, message: str) -> None:
    """Raise RuntimeError with *message* if the release's phase forbids *action*."""
    if not can(release, action):
        raise RuntimeError(message)


def new_id(prefix: str) -> str:
    return f"{prefix}_{uuid.uuid4().hex[:12]}"


def normalize_name(value: str | None) -> str:
    text = (value or "").strip().lower()
    text = re.sub(r"\s+", "", text)
    return text.replace("_", "-").replace(".", "-")


def split_list(value: str | None) -> list[str]:
    if not value:
        return []
    parts = re.split(r"[,，、;；/]+", value)
    return sorted({part.strip() for part in parts if part.strip()})


def join_list(values: list[str] | set[str] | tuple[str, ...]) -> str:
    return ",".join(sorted({str(v).strip() for v in values if str(v).strip()}))


def order_chips(values: str | list[str] | set[str] | tuple[str, ...] | None) -> list[str]:
    """Order chip names alphabetically but always keep x201 last.

    Accepts a comma-separated string or any iterable; dedupes and applies the
    x201-last rule used in the app workbench and the QA release report.
    """
    if isinstance(values, str):
        items: list[Any] = re.split(r"[,，、;；/]+", values)
    else:
        items = list(values or [])
    seen: list[str] = []
    for item in items:
        text = str(item).strip()
        if text and text not in seen:
            seen.append(text)
    rest = sorted((c for c in seen if c.lower() != "x201"), key=str.lower)
    tail = [c for c in seen if c.lower() == "x201"]
    return rest + tail


def loads_json(value: str | None, default: Any) -> Any:
    if not value:
        return default
    return json.loads(value)


def dumps_json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True)


def connect(path: str | Path = "release_system.db") -> sqlite3.Connection:
    conn = sqlite3.connect(path, timeout=10.0, factory=ManagedConnection)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    conn.execute("PRAGMA busy_timeout = 5000")
    key = str(Path(path).resolve()) if not str(path).startswith(":") else str(path)
    with _INIT_LOCK:
        if key not in _INITIALIZED_DBS:
            try:
                conn.execute("PRAGMA journal_mode = WAL")
            except sqlite3.OperationalError:
                pass  # in-memory dbs don't support WAL
            init_db(conn)
            _INITIALIZED_DBS.add(key)
    return conn


def reset_init_state() -> None:
    """Reset the initialized-db tracker; used by tests when cycling DBs."""
    with _INIT_LOCK:
        _INITIALIZED_DBS.clear()


def backup_sqlite(src_path: str | Path, dest_path: str | Path) -> None:
    """Write a consistent backup of the SQLite db at *src_path* to *dest_path*.

    Uses the SQLite online-backup API rather than a file copy: in WAL mode a
    plain file copy can miss committed transactions still in the -wal file and
    skips the -wal/-shm sidecars, producing a silently stale backup.
    """
    source = sqlite3.connect(src_path)
    try:
        dest = sqlite3.connect(dest_path)
        try:
            source.backup(dest)
        finally:
            dest.close()
    finally:
        source.close()


def init_db(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS apps (
            id TEXT PRIMARY KEY,
            git_url TEXT NOT NULL DEFAULT '',
            git_branch TEXT NOT NULL DEFAULT '',
            aliases_json TEXT NOT NULL DEFAULT '[]',
            created_by TEXT NOT NULL DEFAULT 'import',
            created_at TEXT NOT NULL DEFAULT ''
        );

        CREATE TABLE IF NOT EXISTS releases (
            id TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            maca_version TEXT NOT NULL DEFAULT '',
            app_freeze_deadline TEXT NOT NULL DEFAULT '',
            doc_deadline TEXT NOT NULL DEFAULT '',
            released_locked INTEGER NOT NULL DEFAULT 0,
            released_locked_at TEXT NOT NULL DEFAULT '',
            released_locked_by TEXT NOT NULL DEFAULT '',
            created_at TEXT NOT NULL,
            source TEXT NOT NULL,
            cloned_from TEXT NOT NULL DEFAULT ''
        );

        CREATE TABLE IF NOT EXISTS snapshots (
            release_id TEXT NOT NULL REFERENCES releases(id) ON DELETE CASCADE,
            app_id TEXT NOT NULL REFERENCES apps(id) ON DELETE CASCADE,
            data_json TEXT NOT NULL,
            PRIMARY KEY (release_id, app_id)
        );

        CREATE TABLE IF NOT EXISTS artifacts (
            release_id TEXT NOT NULL REFERENCES releases(id) ON DELETE CASCADE,
            kind TEXT NOT NULL,
            name TEXT NOT NULL,
            content TEXT NOT NULL,
            final INTEGER NOT NULL DEFAULT 0,
            generated_at TEXT NOT NULL,
            PRIMARY KEY (release_id, kind)
        );

        CREATE TABLE IF NOT EXISTS qa_logs (
            release_id TEXT PRIMARY KEY REFERENCES releases(id) ON DELETE CASCADE,
            filename TEXT NOT NULL,
            storage_path TEXT NOT NULL,
            uploaded_at TEXT NOT NULL,
            uploaded_by TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS audit (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ts TEXT NOT NULL,
            user TEXT NOT NULL,
            role TEXT NOT NULL,
            app_id TEXT NOT NULL DEFAULT '',
            release_id TEXT NOT NULL DEFAULT '',
            event TEXT NOT NULL DEFAULT '',
            message TEXT NOT NULL,
            detail TEXT NOT NULL DEFAULT ''
        );

        CREATE TABLE IF NOT EXISTS users (
            username TEXT PRIMARY KEY,
            password_hash TEXT NOT NULL DEFAULT '',
            role TEXT NOT NULL,
            auth_source TEXT NOT NULL DEFAULT 'local',
            display_name TEXT NOT NULL DEFAULT ''
        );

        CREATE TABLE IF NOT EXISTS sessions (
            token TEXT PRIMARY KEY,
            username TEXT NOT NULL REFERENCES users(username) ON DELETE CASCADE,
            created_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS release_schedule (
            id TEXT PRIMARY KEY,
            version TEXT NOT NULL,
            branch_cut_at TEXT NOT NULL DEFAULT '',
            release_at TEXT NOT NULL DEFAULT '',
            note TEXT NOT NULL DEFAULT '',
            created_at TEXT NOT NULL,
            created_by TEXT NOT NULL,
            updated_at TEXT NOT NULL DEFAULT '',
            updated_by TEXT NOT NULL DEFAULT ''
        );

        CREATE TABLE IF NOT EXISTS cicd_tasks (
            id TEXT PRIMARY KEY,
            app_name TEXT NOT NULL,
            app_version TEXT NOT NULL DEFAULT '',
            repo_type TEXT NOT NULL DEFAULT 'git',
            repo_name TEXT NOT NULL DEFAULT '',
            branch TEXT NOT NULL DEFAULT '',
            build_product TEXT NOT NULL DEFAULT '[]',
            community_artifact TEXT NOT NULL DEFAULT '[]',
            build_image TEXT NOT NULL DEFAULT '',
            test_timeout INTEGER NOT NULL DEFAULT 40,
            supports_maca_hpcc TEXT NOT NULL DEFAULT 'No',
            owner_username TEXT NOT NULL,
            status TEXT NOT NULL DEFAULT 'Running',
            notes TEXT NOT NULL DEFAULT '',
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS cicd_task_requests (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            task_id TEXT,
            request_type TEXT NOT NULL DEFAULT 'create',
            payload TEXT NOT NULL DEFAULT '{}',
            submitter TEXT NOT NULL,
            submitter_display TEXT NOT NULL DEFAULT '',
            submitted_at TEXT NOT NULL,
            status TEXT NOT NULL DEFAULT 'pending',
            reviewer TEXT NOT NULL DEFAULT '',
            reviewed_at TEXT NOT NULL DEFAULT '',
            review_note TEXT NOT NULL DEFAULT '',
            is_self_approved INTEGER NOT NULL DEFAULT 0
        );

        CREATE TABLE IF NOT EXISTS cicd_notifications (
            username TEXT NOT NULL,
            last_visited_at TEXT NOT NULL DEFAULT '',
            PRIMARY KEY (username)
        );
        """
    )
    ensure_default_user(conn)
    # --- online migration: add columns introduced after initial deployment ---
    for _col, _col_def in [
        ("auth_source", "TEXT NOT NULL DEFAULT 'local'"),
        ("display_name", "TEXT NOT NULL DEFAULT ''"),
    ]:
        try:
            conn.execute(f"ALTER TABLE users ADD COLUMN {_col} {_col_def}")
        except sqlite3.OperationalError:
            pass  # column already exists
    for _tbl, _col, _col_def in [
        ("cicd_task_requests", "approval_mode",    "TEXT NOT NULL DEFAULT 'immediate'"),
        ("cicd_task_requests", "delivery_status",  "TEXT NOT NULL DEFAULT ''"),
        ("cicd_task_requests", "jira_id",           "TEXT NOT NULL DEFAULT ''"),
        ("cicd_task_requests", "jira_auto_created", "INTEGER NOT NULL DEFAULT 0"),
        ("cicd_task_requests", "delivered_by",     "TEXT NOT NULL DEFAULT ''"),
        ("cicd_task_requests", "delivered_at",     "TEXT NOT NULL DEFAULT ''"),
        ("cicd_task_requests", "returned_reason",  "TEXT NOT NULL DEFAULT ''"),
        ("cicd_task_requests", "returned_at",      "TEXT NOT NULL DEFAULT ''"),
        ("cicd_tasks",         "community_artifact", "TEXT NOT NULL DEFAULT '[]'"),
    ]:
        try:
            conn.execute(f"ALTER TABLE {_tbl} ADD COLUMN {_col} {_col_def}")
        except sqlite3.OperationalError:
            pass
    conn.commit()


def hash_password(password: str, salt: str | None = None) -> str:
    salt = salt or secrets.token_hex(16)
    digest = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt.encode("utf-8"), 120_000).hex()
    return f"{salt}${digest}"


def verify_password(password: str, encoded: str) -> bool:
    salt, expected = encoded.split("$", 1)
    return secrets.compare_digest(hash_password(password, salt).split("$", 1)[1], expected)


DEFAULT_USERS: tuple[tuple[str, str, str], ...] = (
    ("rm", "rm", "RM"),
    ("owner_test", "owner_test", "Owner"),
    ("qa", "qa", "QA"),
    ("spd_test", "spd_test", "SPD"),
)

ROLES = {"RM", "Owner", "QA", "Admin", "SPD"}
RELEASE_DECISIONS = {"release", "cicd_only", "stopped"}
NON_RELEASE_DECISIONS = {"cicd_only", "stopped"}
DOC_TARGETS = {"manual", "ai4sci"}
QA_STATUSES = {"not_checked", "qa_passed", "has_issues", "cannot_release"}
MAX_APP_DESCRIPTION_CHARS = 30
MANAGER_REVIEW_FIELDS = [
    ("app_name", "App"),
    ("official_name", "官方名称"),
    ("doc_target", "文档类型"),
    ("app_type", "App类型"),
    ("version", "版本号"),
    ("owners", "Owner"),
    ("chip_support", "支持芯片类型"),
    ("x86_chips", "X86支持芯片"),
    ("arm_chips", "ARM支持芯片"),
    ("release_decision", "Release决策"),
    ("qa_status", "QA状态"),
    ("owner_confirmed", "Owner确认"),
    ("releasable", "是否可发布"),
    ("not_releasable_reason", "不可发布原因"),
    ("known_limitations", "已知限制"),
    ("gerrit_url", "Gerrit URL"),
    ("git_branch", "Branch"),
]
DEFAULT_MANAGER_REVIEW_FIELDS = [
    "app_name",
    "version",
    "owners",
    "chip_support",
    "releasable",
    "not_releasable_reason",
    "known_limitations",
]


def normalize_release_decision(value: str | None) -> str:
    decision = (value or "release").strip()
    return "stopped" if decision == "no_release" else decision


def normalize_doc_target(value: str | None) -> str:
    target = (value or "manual").strip()
    aliases = {
        "HPC": "manual",
        "hpc": "manual",
        "manual": "manual",
        "AI4Sci": "ai4sci",
        "ai4sci": "ai4sci",
        "AI4SCI": "ai4sci",
    }
    return aliases.get(target, "manual")


def app_description_count(value: str | None) -> int:
    text = (value or "").strip()
    count = 0
    i = 0
    while i < len(text):
        ch = text[i]
        if ch.isspace():
            i += 1
            continue
        if ch.isascii() and ch.isalnum():
            while i < len(text) and text[i].isascii() and text[i].isalnum():
                i += 1
            count += 1
            continue
        count += 1
        i += 1
    return count


def normalize_app_description(value: str | None) -> str:
    description = (value or "").strip()
    if app_description_count(description) > MAX_APP_DESCRIPTION_CHARS:
        raise ValueError(f"描述不能超过{MAX_APP_DESCRIPTION_CHARS}字")
    return description


def ensure_default_user(conn: sqlite3.Connection) -> None:
    for username, password, role in DEFAULT_USERS:
        conn.execute(
            "INSERT INTO users(username, password_hash, role) VALUES (?, ?, ?) ON CONFLICT(username) DO NOTHING",
            (username, hash_password(password), role),
        )


def create_user(conn: sqlite3.Connection, username: str, password: str, role: str = "Owner") -> None:
    with transaction(conn):
        conn.execute(
            "INSERT INTO users(username, password_hash, role) VALUES (?, ?, ?) ON CONFLICT(username) DO UPDATE SET password_hash=excluded.password_hash, role=excluded.role",
            (username, hash_password(password), role),
        )


def clear_business_data(conn: sqlite3.Connection, *, user: str = "admin", role: str = "Admin") -> None:
    """Clear release data while preserving user accounts."""
    with transaction(conn):
        conn.execute("DELETE FROM artifacts")
        conn.execute("DELETE FROM qa_logs")
        conn.execute("DELETE FROM snapshots")
        conn.execute("DELETE FROM releases")
        conn.execute("DELETE FROM apps")
        conn.execute("DELETE FROM release_schedule")
        conn.execute("DELETE FROM audit")
        ensure_default_user(conn)
        audit(conn, "数据库已清空，默认账号已保留", user=user, role=role)


def authenticate(conn: sqlite3.Connection, username: str, password: str) -> str | None:
    row = conn.execute("SELECT username, password_hash FROM users WHERE username = ?", (username,)).fetchone()
    if not row or not verify_password(password, row["password_hash"]):
        return None
    token = secrets.token_urlsafe(32)
    with transaction(conn):
        conn.execute("INSERT INTO sessions(token, username, created_at) VALUES (?, ?, ?)", (token, username, now()))
    return token


def session_user(conn: sqlite3.Connection, token: str | None) -> dict[str, str] | None:
    if not token:
        return None
    row = conn.execute(
        "SELECT users.username, users.role FROM sessions JOIN users ON sessions.username = users.username WHERE sessions.token = ?",
        (token,),
    ).fetchone()
    return dict(row) if row else None


def logout_session(conn: sqlite3.Connection, token: str | None) -> None:
    if token:
        with transaction(conn):
            conn.execute("DELETE FROM sessions WHERE token = ?", (token,))


def ldap_login_or_create(conn: sqlite3.Connection, username: str, display_name: str = "") -> str:
    """Ensure an LDAP-authenticated user exists in the local users table, then return a session token.

    First-time LDAP logins are auto-provisioned with role='Owner'.  Subsequent logins
    update display_name if it changed; the stored role is always preserved.
    """
    with transaction(conn):
        row = conn.execute(
            "SELECT username, role, display_name FROM users WHERE username = ?", (username,)
        ).fetchone()
        if not row:
            conn.execute(
                "INSERT INTO users(username, password_hash, role, auth_source, display_name) "
                "VALUES (?, '', 'Owner', 'ldap', ?)",
                (username, display_name),
            )
            audit(
                conn,
                f"域账号首次登录，自动创建 Owner 用户：{username}",
                user=username,
                role="Owner",
                event="ldap_first_login",
            )
        elif display_name and dict(row).get("display_name", "") != display_name:
            conn.execute(
                "UPDATE users SET display_name = ? WHERE username = ?", (display_name, username)
            )
    token = secrets.token_urlsafe(32)
    with transaction(conn):
        conn.execute(
            "INSERT INTO sessions(token, username, created_at) VALUES (?, ?, ?)",
            (token, username, now()),
        )
    return token


def list_users(conn: sqlite3.Connection) -> list[dict[str, Any]]:
    """Return all users ordered by auth_source then username."""
    return [
        dict(row)
        for row in conn.execute(
            "SELECT username, role, auth_source, display_name FROM users ORDER BY auth_source, role, username"
        )
    ]


def set_user_role(
    conn: sqlite3.Connection,
    username: str,
    new_role: str,
    *,
    actor: str,
    actor_role: str,
) -> None:
    """Update a user's role; actor must be Admin."""
    if new_role not in ROLES:
        raise ValueError(f"无效角色：{new_role}，合法值为 {sorted(ROLES)}")
    row = conn.execute("SELECT role FROM users WHERE username = ?", (username,)).fetchone()
    if not row:
        raise KeyError(f"用户不存在：{username}")
    old_role = dict(row)["role"]
    with transaction(conn):
        conn.execute("UPDATE users SET role = ? WHERE username = ?", (new_role, username))
        audit(
            conn,
            f"修改用户角色：{username}  {old_role} → {new_role}",
            user=actor,
            role=actor_role,
            event="set_user_role",
            detail=[{"field": "role", "label": "角色", "old": old_role, "new": new_role}],
        )


def audit(
    conn: sqlite3.Connection,
    message: str,
    *,
    user: str = "system",
    role: str = "system",
    app_id: str = "",
    release_id: str = "",
    event: str = "",
    detail: Any = "",
) -> None:
    detail_text = detail if isinstance(detail, str) else dumps_json(detail)
    with transaction(conn):
        conn.execute(
            "INSERT INTO audit(ts, user, role, app_id, release_id, event, message, detail) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            (now(), user, role, app_id, release_id, event, message, detail_text),
        )


def app_audit_log(conn: sqlite3.Connection, app_id: str, release_id: str = "") -> list[dict[str, Any]]:
    """Audit entries for an app. If *release_id* is given, only that release's."""
    cols = "SELECT ts, user, role, release_id, event, message, detail FROM audit WHERE app_id = ?"
    if release_id:
        rows = conn.execute(cols + " AND release_id = ? ORDER BY id DESC", (app_id, release_id))
    else:
        rows = conn.execute(cols + " ORDER BY id DESC", (app_id,))
    entries = []
    for row in rows:
        entry = dict(row)
        entry["detail"] = loads_json(entry.get("detail"), []) if entry.get("detail") else []
        entries.append(entry)
    return entries


def fmt_audit_value(v: Any) -> str:
    """Render an audit old/new value as a display string."""
    if v is None:
        return ""
    if isinstance(v, (list, tuple)):
        return ", ".join(str(x) for x in v)
    if isinstance(v, bool):
        return "是" if v else "否"
    return str(v)


def field_diff(before: dict[str, Any], after: dict[str, Any], labels: dict[str, str]) -> list[dict[str, str]]:
    """Return [{field,label,old,new}] for keys in *labels* whose value changed."""
    changes: list[dict[str, str]] = []
    for key, label in labels.items():
        old = before.get(key)
        new = after.get(key)
        if old == new:
            continue
        changes.append({"field": key, "label": label, "old": fmt_audit_value(old), "new": fmt_audit_value(new)})
    return changes


TEST_DOC_FIELD_LABELS = {
    "command": "命令",
    "dataset": "测试数据集",
    "content": "测试内容",
    "result_view": "结果查看",
    "pass_criteria": "通过标准",
}


def test_docs_diff(before: list[dict[str, Any]], after: list[dict[str, Any]]) -> list[dict[str, str]]:
    """Return field-level [{field,label,old,new}] entries between two test-doc lists."""
    by_id = {d.get("id"): d for d in before}
    changes: list[dict[str, str]] = []
    for doc in after:
        path = doc.get("path") or doc.get("id") or "test"
        old = by_id.get(doc.get("id"))
        if old is None:
            changes.append({"field": str(doc.get("id")), "label": f"{path}（新增测试项）", "old": "", "new": "已添加"})
            continue
        for key, label in TEST_DOC_FIELD_LABELS.items():
            if (old.get(key) or "") != (doc.get(key) or ""):
                changes.append({"field": f"{path}.{key}", "label": f"{path} · {label}",
                                "old": fmt_audit_value(old.get(key)), "new": fmt_audit_value(doc.get(key))})
    return changes


def read_csv(path: str | Path) -> list[dict[str, str]]:
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        return [{k: (v or "").strip() for k, v in row.items()} for row in csv.DictReader(f)]


def parse_csv_text(text: str) -> list[dict[str, str]]:
    return [{k: (v or "").strip() for k, v in row.items()} for row in csv.DictReader(io.StringIO(text.lstrip("﻿")))]


def parse_alias_lines(raw: str = "") -> dict[str, str]:
    aliases: dict[str, str] = {}
    for line in raw.splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        left, right = [part.strip() for part in line.split("=", 1)]
        if left and right:
            aliases[normalize_name(left)] = normalize_name(right)
    return aliases


def infer_doc_target(category: str = "", app_type: str = "") -> str:
    """Pick the documentation target. AI-for-Science apps document into the
    AI4Sci guide; HPC apps and 工具 document into the HPC manual.

    类别 (category) is the authoritative signal. app_type is only a fallback
    for CSVs without a 类别 column — and an HPC app_type such as
    'HPC框架/工具' must NOT be misread as AI4Sci just for containing '框架'.
    """
    cat = str(category or "").strip().lower()
    typ = str(app_type or "").strip().lower()
    if "ai for science" in cat or "ai4sci" in cat:
        return "ai4sci"
    if "hpc" in cat or "工具" in cat:
        return "manual"
    # category absent / unrecognized — fall back to app_type markers
    if "ai" in typ or typ.endswith("模型"):
        return "ai4sci"
    return "manual"


def csv_value(row: dict[str, str], *names: str) -> str:
    for name in names:
        value = (row.get(name) or "").strip()
        if value:
            return value
    return ""


def csv_checkmark(value: str | None) -> bool:
    """Interpret a CSV sanity cell as a boolean pass mark.

    '✔' / '√' etc. (and a few affirmative words) count as passed; empty
    cells or descriptive notes like 'arm sanity' do not.
    """
    v = str(value or "").strip().lower()
    if not v:
        return False
    if any(mark in v for mark in ("✔", "✓", "√", "✅")):
        return True
    return v in ("pass", "passed", "ok", "yes", "y", "true", "1", "通过", "已通过")


def init_csv_official_name(row: dict[str, str]) -> str:
    return csv_value(row, "官方名称", "名称", "app_name")


def init_csv_doc_category(row: dict[str, str]) -> str:
    if "APP类型" in row:
        return csv_value(row, "类型")
    return csv_value(row, "类别", "类型")


def init_csv_app_type(row: dict[str, str]) -> str:
    return csv_value(row, "APP类型", "类型")


def canonical_id(name: str, aliases: dict[str, str] | None = None) -> str:
    normalized = normalize_name(name)
    return (aliases or {}).get(normalized, normalized)


SNAPSHOT_META_FIELDS = ("official_name", "type", "official_url", "description", "doc_target", "owners")
APP_META_LABELS = {
    "official_name": "官方名称",
    "type": "App类型",
    "official_url": "官方 URL",
    "description": "描述",
    "doc_target": "文档类型",
    "owners": "Owner",
}


def display_name(official_name: str | None, version: str | None = "") -> str:
    """Human-facing app name: official name plus version when known."""
    official = (official_name or "").strip()
    ver = (version or "").strip()
    return f"{official} {ver}".strip() if ver else official


def app_view(app: dict[str, Any], snapshot: dict[str, Any] | None) -> dict[str, Any]:
    """Merge the global app row with per-release snapshot metadata.

    official_name/type/official_url/description/doc_target/owners live on the
    snapshot (per-release); id/git_url/git_branch are global app identity.
    """
    snapshot = snapshot or {}
    view = {
        "id": app.get("id", ""),
        "git_url": app.get("git_url", ""),
        "git_branch": app.get("git_branch", ""),
        "aliases": app.get("aliases", []),
        "created_by": app.get("created_by", ""),
        "created_at": app.get("created_at", ""),
        "official_name": snapshot.get("official_name", ""),
        "type": snapshot.get("type", ""),
        "official_url": snapshot.get("official_url", ""),
        "description": snapshot.get("description", ""),
        "doc_target": normalize_doc_target(snapshot.get("doc_target")),
        "owners": list(snapshot.get("owners", []) or []),
        "version": snapshot.get("version", ""),
    }
    view["name"] = display_name(view["official_name"], view["version"])
    return view


def variant_app_id(base_id: str, version: str, branch: str, used_ids: set[str]) -> str:
    suffix = normalize_name(version) or normalize_name(branch) or "variant"
    candidate = f"{base_id}_{suffix}" if suffix else base_id
    branch_suffix = normalize_name(branch)
    if candidate in used_ids and branch_suffix and branch_suffix not in candidate:
        candidate = f"{candidate}_{branch_suffix}"
    index = 2
    original = candidate
    while candidate in used_ids:
        candidate = f"{original}_{index}"
        index += 1
    return candidate


def row_to_app(row: sqlite3.Row) -> dict[str, Any]:
    data = dict(row)
    data["aliases"] = loads_json(data.pop("aliases_json"), [])
    return data


def save_app(conn: sqlite3.Connection, app: dict[str, Any]) -> None:
    conn.execute(
        """
        INSERT INTO apps(id, git_url, git_branch, aliases_json, created_by, created_at)
        VALUES (?, ?, ?, ?, ?, ?)
        ON CONFLICT(id) DO UPDATE SET
          git_url=excluded.git_url,
          git_branch=excluded.git_branch,
          aliases_json=excluded.aliases_json,
          created_by=excluded.created_by
        """,
        (
            app["id"],
            app.get("git_url", ""),
            app.get("git_branch", ""),
            dumps_json(sorted(set(app.get("aliases", [])))),
            app.get("created_by", "import"),
            app.get("created_at") or now(),
        ),
    )


def get_app(conn: sqlite3.Connection, app_id: str) -> dict[str, Any]:
    row = conn.execute("SELECT * FROM apps WHERE id = ?", (app_id,)).fetchone()
    if not row:
        raise KeyError(f"Unknown app: {app_id}")
    return row_to_app(row)


def list_apps(conn: sqlite3.Connection) -> list[dict[str, Any]]:
    return [row_to_app(row) for row in conn.execute("SELECT * FROM apps ORDER BY id")]


def locked_releases_for_app(conn: sqlite3.Connection, app_id: str) -> list[str]:
    return [
        row["name"]
        for row in conn.execute(
            """
            SELECT releases.name
            FROM releases
            JOIN snapshots ON snapshots.release_id = releases.id
            WHERE snapshots.app_id = ? AND releases.released_locked = 1
            ORDER BY releases.created_at
            """,
            (app_id,),
        )
    ]


def delete_app(conn: sqlite3.Connection, app_id: str, *, user: str = "admin", role: str = "Admin") -> dict[str, Any]:
    app = get_app(conn, app_id)
    locked_releases = locked_releases_for_app(conn, app_id)
    if locked_releases:
        raise RuntimeError(f"App is used by locked releases and cannot be deleted: {', '.join(locked_releases)}")
    with transaction(conn):
        affected_releases = [
            row["release_id"]
            for row in conn.execute("SELECT release_id FROM snapshots WHERE app_id = ?", (app_id,))
        ]
        for affected in affected_releases:
            conn.execute("DELETE FROM artifacts WHERE release_id = ? AND final = 0", (affected,))
        conn.execute("DELETE FROM apps WHERE id = ?", (app_id,))
        audit(conn, f"删除 app：{app_id}", user=user, role=role, app_id=app_id, event="delete_app")
    return app


def base_snapshot(
    app_id: str,
    *,
    official_name: str = "",
    app_type: str = "",
    official_url: str = "",
    description: str = "",
    doc_target: str = "manual",
    owners: list[str] | None = None,
) -> dict[str, Any]:
    return {
        "app_id": app_id,
        "official_name": official_name,
        "type": app_type,
        "official_url": official_url,
        "description": description,
        "doc_target": normalize_doc_target(doc_target),
        "owners": sorted(set(owners or [])),
        "release_decision": "release",
        "qa_status": "not_checked",
        "qa_issue_note": "",
        "owner_confirmed": False,
        "version": "",
        "x86_chips": "",
        "arm_chips": "",
        "hpcc_chip": "",
        "arch": "",
        "python_labels": "",
        "pytorch_labels": "",
        "build_os": "",
        "build_arches": "",
        "maca_version": "",
        "doc": {
            "intro": "",
            "image_usage": "",
            "binary_usage": "",
            "env_setup": "",
            "limitations": "",
        },
        "community": {
            "release_status": "",
            "python_version": "",
            "framework_version": "",
        },
        "sanity": {
            "arm_kylin": False,
            "ubuntu": False,
        },
        "app_info": None,
        "app_info_diffs": [],
        "test_docs": [],
        "missing_items": [],
    }


def save_release(conn: sqlite3.Connection, release: dict[str, Any]) -> None:
    conn.execute(
        """
        INSERT INTO releases(id, name, maca_version, app_freeze_deadline, doc_deadline,
                             released_locked, released_locked_at, released_locked_by,
                             created_at, source, cloned_from)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(id) DO UPDATE SET
          name=excluded.name,
          maca_version=excluded.maca_version,
          app_freeze_deadline=excluded.app_freeze_deadline,
          doc_deadline=excluded.doc_deadline,
          released_locked=excluded.released_locked,
          released_locked_at=excluded.released_locked_at,
          released_locked_by=excluded.released_locked_by,
          created_at=excluded.created_at,
          source=excluded.source,
          cloned_from=excluded.cloned_from
        """,
        (
            release["id"],
            release["name"],
            release.get("maca_version", ""),
            normalize_deadline(release.get("app_freeze_deadline", "")),
            normalize_deadline(release.get("doc_deadline", "")),
            int(release.get("released_locked", 0)),
            release.get("released_locked_at", ""),
            release.get("released_locked_by", ""),
            release.get("created_at", now()),
            release.get("source", "manual"),
            release.get("cloned_from", ""),
        ),
    )


def update_release_deadlines(
    conn: sqlite3.Connection,
    release_id: str,
    *,
    name: str | None = None,
    app_freeze_deadline: str | None = None,
    doc_deadline: str | None = None,
    user: str = "system",
    role: str = "system",
) -> dict[str, Any]:
    release = get_release(conn, release_id)
    if release.get("released_locked"):
        raise RuntimeError("Release 已最终锁定，不可修改 release 设置")
    new_name = (name or "").strip() if name is not None else release.get("name", "")
    if not new_name:
        raise ValueError("Release 名称不能为空")
    new_freeze = normalize_deadline(app_freeze_deadline) if app_freeze_deadline is not None else release.get("app_freeze_deadline", "")
    new_doc = normalize_deadline(doc_deadline) if doc_deadline is not None else release.get("doc_deadline", "")
    validate_deadline_order(new_freeze, new_doc)
    with transaction(conn):
        conn.execute(
            "UPDATE releases SET name = ?, app_freeze_deadline = ?, doc_deadline = ? WHERE id = ?",
            (new_name, new_freeze, new_doc, release_id),
        )
        audit(
            conn,
            f"更新 release 设置：{release['name']} -> {new_name}，app_freeze={new_freeze or '空'}, doc={new_doc or '空'}",
            user=user,
            role=role,
            release_id=release_id,
            event="update_release_settings",
        )
    return get_release(conn, release_id)


def normalize_schedule_date(value: str | None) -> str:
    """Normalize a schedule date string to ``YYYY-MM-DD``.

    Schedule entries store calendar dates only (no time-of-day); strip any
    trailing time component so the table renders cleanly.
    """
    text = (value or "").strip()
    if not text:
        return ""
    text = text.replace("T", " ").split(" ")[0]
    try:
        return dt.datetime.strptime(text, "%Y-%m-%d").strftime("%Y-%m-%d")
    except ValueError as exc:
        raise ValueError(f"Invalid date: {value!r}; expected YYYY-MM-DD") from exc


def list_release_schedule(conn: sqlite3.Connection) -> list[dict[str, Any]]:
    rows = conn.execute(
        "SELECT * FROM release_schedule "
        "ORDER BY CASE WHEN branch_cut_at = '' THEN 1 ELSE 0 END, branch_cut_at, "
        "         CASE WHEN release_at = '' THEN 1 ELSE 0 END, release_at, created_at"
    )
    return [dict(row) for row in rows]


def upsert_release_schedule(
    conn: sqlite3.Connection,
    *,
    entry_id: str | None,
    version: str,
    branch_cut_at: str,
    release_at: str,
    note: str = "",
    user: str = "system",
    role: str = "system",
) -> dict[str, Any]:
    version = (version or "").strip()
    if not version:
        raise ValueError("版本号不能为空")
    branch_cut = normalize_schedule_date(branch_cut_at)
    release_date = normalize_schedule_date(release_at)
    if branch_cut and release_date and branch_cut > release_date:
        raise ValueError("拉 branch 时间不能晚于发布时间")
    note = (note or "").strip()
    with transaction(conn):
        existing = None
        if entry_id:
            row = conn.execute("SELECT * FROM release_schedule WHERE id = ?", (entry_id,)).fetchone()
            existing = dict(row) if row else None
        if existing:
            conn.execute(
                "UPDATE release_schedule SET version = ?, branch_cut_at = ?, release_at = ?, note = ?, "
                "updated_at = ?, updated_by = ? WHERE id = ?",
                (version, branch_cut, release_date, note, now(), user, entry_id),
            )
            audit(
                conn,
                f"更新发布时间线：{version}",
                user=user,
                role=role,
                event="update_release_schedule",
                detail=field_diff(
                    existing,
                    {"version": version, "branch_cut_at": branch_cut, "release_at": release_date, "note": note},
                    {"version": "版本号", "branch_cut_at": "拉 branch 时间", "release_at": "发布时间", "note": "备注"},
                ),
            )
            final_id = entry_id
        else:
            final_id = entry_id or new_id("sched")
            conn.execute(
                "INSERT INTO release_schedule(id, version, branch_cut_at, release_at, note, "
                "created_at, created_by, updated_at, updated_by) VALUES (?, ?, ?, ?, ?, ?, ?, '', '')",
                (final_id, version, branch_cut, release_date, note, now(), user),
            )
            audit(
                conn,
                f"新增发布时间线：{version}",
                user=user,
                role=role,
                event="create_release_schedule",
                detail={"version": version, "branch_cut_at": branch_cut, "release_at": release_date, "note": note},
            )
        row = conn.execute("SELECT * FROM release_schedule WHERE id = ?", (final_id,)).fetchone()
        return dict(row)


def delete_release_schedule(
    conn: sqlite3.Connection,
    entry_id: str,
    *,
    user: str = "system",
    role: str = "system",
) -> bool:
    with transaction(conn):
        row = conn.execute("SELECT version FROM release_schedule WHERE id = ?", (entry_id,)).fetchone()
        if not row:
            return False
        conn.execute("DELETE FROM release_schedule WHERE id = ?", (entry_id,))
        audit(
            conn,
            f"删除发布时间线：{row['version']}",
            user=user,
            role=role,
            event="delete_release_schedule",
            detail={"id": entry_id, "version": row["version"]},
        )
        return True


def release_is_locked(conn: sqlite3.Connection, release_id: str) -> bool:
    row = conn.execute("SELECT released_locked FROM releases WHERE id = ?", (release_id,)).fetchone()
    return bool(row and row["released_locked"])


def save_snapshot(conn: sqlite3.Connection, release_id: str, app_id: str, snapshot: dict[str, Any]) -> None:
    if release_is_locked(conn, release_id):
        raise RuntimeError("Release 已最终锁定，所有快照不可修改")
    conn.execute(
        """
        INSERT INTO snapshots(release_id, app_id, data_json)
        VALUES (?, ?, ?)
        ON CONFLICT(release_id, app_id) DO UPDATE SET data_json=excluded.data_json
        """,
        (release_id, app_id, dumps_json(snapshot)),
    )


def get_release(conn: sqlite3.Connection, release_id: str) -> dict[str, Any]:
    row = conn.execute("SELECT * FROM releases WHERE id = ?", (release_id,)).fetchone()
    if not row:
        raise KeyError(f"Unknown release: {release_id}")
    release = dict(row)
    release["released_locked"] = bool(release.get("released_locked"))
    release["snapshots"] = {
        snap["app_id"]: loads_json(snap["data_json"], {})
        for snap in conn.execute("SELECT app_id, data_json FROM snapshots WHERE release_id = ?", (release_id,))
    }
    release["phase"] = current_phase(release)
    return release


def list_releases(conn: sqlite3.Connection) -> list[dict[str, Any]]:
    result = []
    for row in conn.execute("SELECT * FROM releases ORDER BY created_at, rowid"):
        rel = dict(row)
        rel["released_locked"] = bool(rel.get("released_locked"))
        rel["phase"] = current_phase(rel)
        result.append(rel)
    return result


def previous_release(conn: sqlite3.Connection, release_id: str) -> dict[str, Any] | None:
    releases = list_releases(conn)
    for i, release in enumerate(releases):
        if release["id"] == release_id and i > 0:
            return get_release(conn, releases[i - 1]["id"])
    return None


def last_published_snapshot(conn: sqlite3.Connection, app_id: str, before_release_id: str) -> dict[str, Any] | None:
    """Find the most recent locked snapshot for *app_id* before *before_release_id*."""
    releases = list_releases(conn)
    target_idx = next((i for i, r in enumerate(releases) if r["id"] == before_release_id), None)
    if target_idx is None:
        return None
    for i in range(target_idx - 1, -1, -1):
        r = releases[i]
        if not r.get("released_locked"):
            continue
        row = conn.execute(
            "SELECT data_json FROM snapshots WHERE release_id = ? AND app_id = ?",
            (r["id"], app_id),
        ).fetchone()
        if row:
            snap = loads_json(row["data_json"], {})
            if snap.get("locked_in_release"):
                return snap
    return None


def import_initial(
    conn: sqlite3.Connection,
    csv_path: str | Path,
    *,
    release_name: str | None = None,
    maca_version: str | None = None,
    app_freeze_deadline: str = "",
    doc_deadline: str = "",
) -> str:
    return import_initial_rows(
        conn,
        read_csv(csv_path),
        release_name=release_name,
        maca_version=maca_version,
        app_freeze_deadline=app_freeze_deadline,
        doc_deadline=doc_deadline,
    )


def import_initial_rows(
    conn: sqlite3.Connection,
    rows: list[dict[str, str]],
    *,
    release_name: str | None = None,
    maca_version: str | None = None,
    app_freeze_deadline: str = "",
    doc_deadline: str = "",
) -> str:
    """Import a single init CSV: one app per (git_url, git_branch) pair.

    Supports the legacy release CSV columns plus the rich hpc_app.csv shape:
    类别, id, 名称, Owner, 类型, 描述, git_url, git_branch, 对应官方版本,
    X86支持芯片系列, ARM支持芯片类型, 开发者社区发布*, *sanity.
    Rows sharing a (git_url, git_branch) pair form one app; rows without a
    git repo (e.g. 停止发布 entries) are skipped rather than failing import.
    """
    validate_deadline_order(app_freeze_deadline, doc_deadline)
    if not rows:
        raise ValueError("初始化 CSV 为空")
    if conn.execute("SELECT 1 FROM releases LIMIT 1").fetchone():
        raise RuntimeError("已存在 release，初始化导入只能在空库执行；如需开启新一轮，请使用「克隆 release」")

    groups: dict[tuple[str, str], list[dict[str, str]]] = {}
    order: list[tuple[str, str]] = []
    for row in rows:
        key = ((row.get("git_url") or "").strip(), (row.get("git_branch") or "").strip())
        if not key[0] or not key[1]:
            # no repo -> nothing buildable; skip instead of aborting import
            continue
        if key not in groups:
            groups[key] = []
            order.append(key)
        groups[key].append(row)

    if not order:
        raise ValueError("初始化 CSV 没有可导入的行（每行需要 git_url 和 git_branch）")

    base_counts: dict[str, int] = {}
    for key in order:
        bid = normalize_name(init_csv_official_name(groups[key][0])) or "app"
        base_counts[bid] = base_counts.get(bid, 0) + 1

    used_ids: set[str] = set()
    built: list[tuple[dict[str, Any], dict[str, Any]]] = []
    for key in order:
        group = groups[key]
        first = group[0]
        official = init_csv_official_name(first)
        bid = normalize_name(official) or "app"
        if base_counts[bid] > 1 or bid in used_ids:
            app_id = variant_app_id(bid, first.get("app_version", ""), key[1], used_ids)
        else:
            app_id = bid
        used_ids.add(app_id)

        x86: list[str] = []
        arm: list[str] = []
        hpcc: list[str] = []
        owners: list[str] = []
        for r in group:
            # new hpc_app.csv carries explicit X86 / ARM chip columns;
            # the legacy release CSV splits maca_chip by the arch column.
            x86.extend(split_list(csv_value(r, "X86支持芯片系列")))
            arm.extend(split_list(csv_value(r, "ARM支持芯片类型")))
            arch = (r.get("arch") or "").lower()
            chips = split_list(r.get("maca_chip"))
            (arm if "arm" in arch or "aarch" in arch else x86).extend(chips)
            hpcc.extend(split_list(r.get("hpcc_chip")))
            owners.extend(split_list(r.get("Owner")))

        app = {
            "id": app_id,
            "git_url": key[0],
            "git_branch": key[1],
            "aliases": sorted({official} if official else set()),
            "created_by": "import",
            "created_at": now(),
        }
        snapshot = base_snapshot(
            app_id,
            official_name=official,
            app_type=init_csv_app_type(first),
            description=csv_value(first, "描述"),
            doc_target=infer_doc_target(init_csv_doc_category(first), init_csv_app_type(first)),
            owners=sorted(set(owners)),
        )
        snapshot["version"] = csv_value(first, "对应官方版本", "app_version")
        snapshot["x86_chips"] = ",".join(order_chips(x86))
        snapshot["arm_chips"] = ",".join(order_chips(arm))
        snapshot["hpcc_chip"] = join_list(hpcc)
        snapshot["arch"] = join_list((r.get("arch") or "") for r in group)
        snapshot["maca_version"] = (first.get("maca_version") or "").strip()
        snapshot["community"] = {
            "release_status": csv_value(first, "开发者社区发布情况"),
            "python_version": csv_value(first, "开发者社区发布包支持python版本"),
            "framework_version": csv_value(first, "开发者社区发布包支持的底层框架及版本"),
        }
        snapshot["sanity"] = {
            "arm_kylin": csv_checkmark(csv_value(first, "ARM / Kylin sanity")),
            "ubuntu": csv_checkmark(csv_value(first, "Ubuntu sanity / 兼容性sanity", "Ubuntu / 兼容性 sanity")),
        }
        built.append((app, snapshot))

    release_id = new_id("rel")
    csv_maca = (rows[0].get("maca_version") or "").strip()
    with transaction(conn):
        for app, _ in built:
            save_app(conn, app)
            audit(conn, f"导入 app：{app['id']}", user="import", role="system",
                  app_id=app["id"], release_id=release_id, event="create_app")

        save_release(conn, {
            "id": release_id,
            "name": release_name or maca_version or csv_maca or "initial-release",
            "maca_version": maca_version or csv_maca,
            "app_freeze_deadline": app_freeze_deadline,
            "doc_deadline": doc_deadline,
            "released_locked": 0,
            "released_locked_at": "",
            "released_locked_by": "",
            "created_at": now(),
            "source": "initial_csv",
            "cloned_from": "",
        })
        for app, snapshot in built:
            save_snapshot(conn, release_id, app["id"], snapshot)
        audit(conn, f"首次初始化导入完成：{len(built)} 个 app", release_id=release_id, event="import_initial")
    return release_id


def create_release_from_previous(
    conn: sqlite3.Connection,
    name: str,
    *,
    maca_version: str = "",
    app_freeze_deadline: str = "",
    doc_deadline: str = "",
    user: str = "system",
    role: str = "system",
) -> str:
    if not (name or "").strip():
        raise ValueError("新 Release 名称不能为空")
    validate_deadline_order(app_freeze_deadline, doc_deadline)
    releases = list_releases(conn)
    previous = get_release(conn, releases[-1]["id"]) if releases else None
    release_id = new_id("rel")
    release = {
        "id": release_id,
        "name": name,
        "maca_version": maca_version,
        "app_freeze_deadline": app_freeze_deadline,
        "doc_deadline": doc_deadline,
        "released_locked": 0,
        "released_locked_at": "",
        "released_locked_by": "",
        "created_at": now(),
        "source": "cloned_from_previous" if previous else "empty",
        "cloned_from": previous["id"] if previous else "",
    }
    previous_name = previous["name"] if previous else ""
    summary = f"创建 release「{name}」，从「{previous_name}」克隆" if previous else f"创建 release「{name}」"
    with transaction(conn):
        save_release(conn, release)
        for app in list_apps(conn):
            old = previous["snapshots"].get(app["id"]) if previous else None
            if not old:
                continue
            snapshot = json.loads(json.dumps(old))
            snapshot.pop("locked_in_release", None)
            snapshot.update(
                {
                    "qa_status": "not_checked",
                    "qa_issue_note": "",
                    "missing_items": [],
                }
            )
            save_snapshot(conn, release_id, app["id"], snapshot)
            audit(
                conn,
                f"本 app 随 release 从「{previous_name}」克隆而来，已继承上一版的发布信息",
                user=user,
                role=role,
                app_id=app["id"],
                release_id=release_id,
                event="clone_app",
            )
        audit(conn, summary, user=user, role=role, release_id=release_id, event="create_release")
    return release_id


def _future_unlocked_release_ids(conn: sqlite3.Connection, release_id: str) -> list[str]:
    releases = list_releases(conn)
    start = next((idx for idx, release in enumerate(releases) if release["id"] == release_id), None)
    if start is None:
        raise KeyError(f"Unknown release: {release_id}")
    return [release["id"] for release in releases[start:] if not release.get("released_locked")]


def _initial_snapshot_for_future_release(snapshot: dict[str, Any], target_release: dict[str, Any] | None = None) -> dict[str, Any]:
    future = json.loads(json.dumps(snapshot))
    future.pop("locked_in_release", None)
    future.update(
        {
            "qa_status": "not_checked",
            "qa_issue_note": "",
            "missing_items": [],
        }
    )
    if (
        target_release
        and future.get("release_decision") == "release"
        and not can(target_release, "new_app_release")
    ):
        future["release_decision"] = "cicd_only"
    return future


def add_new_app_request(
    conn: sqlite3.Connection,
    release_id: str,
    *,
    official_name: str,
    git_url: str,
    git_branch: str,
    release_decision: str,
    owner: str,
    doc_target: str = "manual",
) -> str:
    raw_decision = (release_decision or "").strip()
    git_url = (git_url or "").strip()
    git_branch = (git_branch or "").strip()
    if not official_name or not git_url or not git_branch or not raw_decision or not owner:
        raise ValueError("New app requires official_name, git_url, git_branch, release_decision, and submitter owner")
    release_decision = normalize_release_decision(raw_decision)
    if release_decision not in RELEASE_DECISIONS:
        raise ValueError(f"Invalid release_decision: {release_decision}")
    doc_target = normalize_doc_target(doc_target)
    release = get_release(conn, release_id)
    if release.get("released_locked"):
        raise RuntimeError("Release 已最终锁定，不可新增 app")
    intended_action = "new_app_release" if release_decision == "release" else "new_app_non_release"
    if not can(release, intended_action):
        if release_decision == "release":
            raise RuntimeError("已过 app 冻结 deadline，不可再新增以 release 状态进入本期的 app")
        raise RuntimeError("当前阶段不允许新增 app")
    duplicate = conn.execute(
        "SELECT id FROM apps WHERE git_url = ? AND git_branch = ?",
        (git_url, git_branch),
    ).fetchone()
    if duplicate:
        raise RuntimeError(
            f"该 Gerrit URL + branch 已登记为 app（id={duplicate['id']}）："
            f"无需重复新增；如需让它参与本 release，请联系现有 owner 或 RM"
        )
    base_id = normalize_name(official_name)
    if not base_id:
        raise ValueError(f"无法由名称生成有效的 app id：{official_name!r}")
    used_ids = {row["id"] for row in conn.execute("SELECT id FROM apps")}
    app_id = base_id if base_id not in used_ids else variant_app_id(base_id, "", git_branch, used_ids)
    app = {
        "id": app_id,
        "git_url": git_url,
        "git_branch": git_branch,
        "aliases": [official_name],
        "created_by": owner,
        "created_at": now(),
    }
    snapshot = base_snapshot(
        app_id,
        official_name=official_name,
        doc_target=doc_target,
        owners=[owner],
    )
    snapshot["release_decision"] = release_decision
    source_name = release["name"]
    with transaction(conn):
        save_app(conn, app)
        for target_release_id in _future_unlocked_release_ids(conn, release_id):
            target_release = get_release(conn, target_release_id)
            if app_id in target_release["snapshots"]:
                continue
            if target_release_id == release_id:
                save_snapshot(conn, target_release_id, app_id, snapshot)
                audit(
                    conn,
                    f"新增 app：{official_name}（owner={owner}，初始决策={release_decision}）",
                    user=owner, role="Owner", app_id=app_id,
                    release_id=target_release_id, event="create_app",
                )
            else:
                save_snapshot(conn, target_release_id, app_id, _initial_snapshot_for_future_release(snapshot, target_release))
                audit(
                    conn,
                    f"本 app 在「{source_name}」新增后，同步到本 release",
                    user=owner, role="Owner", app_id=app_id,
                    release_id=target_release_id, event="sync_app",
                )
    return app_id


def walk_objects(value: Any, visitor: Callable[[dict[str, Any], list[str]], None], path: list[str] | None = None) -> None:
    path = path or []
    if isinstance(value, dict):
        visitor(value, path)
        for key, child in value.items():
            walk_objects(child, visitor, path + [str(key)])
    elif isinstance(value, list):
        for idx, child in enumerate(value):
            walk_objects(child, visitor, path + [str(idx)])


def parse_app_info(raw: str | dict[str, Any]) -> dict[str, Any]:
    data = json.loads(raw) if isinstance(raw, str) else raw
    x86_chips: set[str] = set()
    arm_chips: set[str] = set()
    python_labels: list[str] = []
    pytorch_labels: list[str] = []
    build_os_list: list[str] = []
    build_arch_list: list[str] = []
    build_targets: list[dict[str, Any]] = []
    test_targets: list[dict[str, Any]] = []
    tests: list[dict[str, Any]] = []

    def _add_unique(target: list[str], value: Any) -> None:
        v = str(value or "").strip()
        if v and v not in target:
            target.append(v)

    for env, cfg in (data.get("app_build") or {}).items():
        if not isinstance(cfg, dict):
            continue
        arch = str(cfg.get("arch") or env)
        chips = cfg.get("supported_chip") if isinstance(cfg.get("supported_chip"), list) else []
        enabled = cfg.get("enabled") is not False
        if enabled:
            target = arm_chips if re.search(r"arm|aarch64", arch, re.I) else x86_chips
            target.update(str(chip).upper() for chip in chips)
            _add_unique(python_labels, cfg.get("python_label"))
            _add_unique(pytorch_labels, cfg.get("pytorch_label"))
            _add_unique(build_os_list, cfg.get("os"))
            _add_unique(build_arch_list, cfg.get("arch"))
        build_targets.append({
            "path": env,
            "arch": arch,
            "chips": chips,
            "enabled": enabled,
            "build_target": cfg.get("build_target", ""),
            "python_label": str(cfg.get("python_label") or "").strip(),
            "pytorch_label": str(cfg.get("pytorch_label") or "").strip(),
            "os": str(cfg.get("os") or "").strip(),
        })

    def visitor(node: dict[str, Any], path: list[str]) -> None:
        if "test_cmd" not in node:
            return
        if node.get("enabled") is False:
            return
        if str(node.get("test_period", "")).strip().lower() == "weekly":
            return
        if node.get("ignore_release"):
            return
        supported = node.get("supported_chip") or {}
        if isinstance(supported, dict):
            chips = list(supported.keys())
            arch_list = sorted({str(v) for values in supported.values() for v in (values if isinstance(values, list) else [values])})
        elif isinstance(supported, list):
            chips = [str(v) for v in supported]
            arch_list = []
        else:
            chips = []
            arch_list = []
        test = {
            "id": ".".join(path),
            "name": path[-1] if path else "test",
            "path": ".".join(path),
            "command": str(node.get("test_cmd") or "").strip(),
            "supported_chips": chips,
            "arch_list": arch_list,
            "enabled": node.get("enabled") is not False,
            "container_args": node.get("container_args", ""),
            "image_target": node.get("img_target", ""),
        }
        tests.append(test)
        test_targets.append(
            {
                "path": test["path"],
                "enabled": test["enabled"],
                "command": test["command"],
                "supported_chips": chips,
                "arch_list": arch_list,
                "container_args": test["container_args"],
                "image_target": test["image_target"],
            }
        )

    walk_objects(data.get("app_test") or {}, visitor)
    return {
        "app_name": data.get("app_name", ""),
        "app_version": data.get("app_version", ""),
        "x86_chips": order_chips(x86_chips),
        "arm_chips": order_chips(arm_chips),
        "python_labels": python_labels,
        "pytorch_labels": pytorch_labels,
        "build_os": build_os_list,
        "build_arches": build_arch_list,
        "build_targets": build_targets,
        "test_targets": test_targets,
        "tests": tests,
        "raw": data,
    }


def diff_app_info(old: dict[str, Any] | None, new: dict[str, Any]) -> list[dict[str, Any]]:
    diffs: list[dict[str, Any]] = []

    def add(diff_type: str, field: str, old_value: Any, new_value: Any, qa_impact: bool = True) -> None:
        if old_value != new_value:
            diffs.append({"id": new_id("diff"), "type": diff_type, "field": field, "old_value": old_value, "new_value": new_value, "qa_impact": qa_impact})

    old = old or {}
    add("版本变化", "app_version", old.get("app_version", ""), new.get("app_version", ""))
    add("X86芯片变化", "x86_chips", old.get("x86_chips", []), new.get("x86_chips", []))
    add("ARM芯片变化", "arm_chips", old.get("arm_chips", []), new.get("arm_chips", []))
    add("Python label 变化", "python_labels", old.get("python_labels", []), new.get("python_labels", []))
    add("PyTorch label 变化", "pytorch_labels", old.get("pytorch_labels", []), new.get("pytorch_labels", []))
    add("OS 变化", "build_os", old.get("build_os", []), new.get("build_os", []))
    add("Arch 变化", "build_arches", old.get("build_arches", []), new.get("build_arches", []))
    add(
        "Build target变化",
        "build_targets",
        [f"{x.get('path')}:{x.get('enabled')}:{x.get('build_target')}" for x in old.get("build_targets", [])],
        [f"{x.get('path')}:{x.get('enabled')}:{x.get('build_target')}" for x in new.get("build_targets", [])],
    )
    add("Test target变化", "test_targets", old.get("test_targets", []), new.get("test_targets", []))
    old_tests = {t["path"]: t["command"] for t in old.get("tests", [])}
    new_tests = {t["path"]: t["command"] for t in new.get("tests", [])}
    for path, cmd in new_tests.items():
        if path not in old_tests:
            add("test_cmd新增", path, "", cmd)
        elif old_tests[path] != cmd:
            add("test_cmd修改", path, old_tests[path], cmd)
    for path, cmd in old_tests.items():
        if path not in new_tests:
            add("test_cmd删除", path, cmd, "")
    return diffs


def ensure_test_docs(snapshot: dict[str, Any], parsed: dict[str, Any], diffs: list[dict[str, Any]]) -> None:
    snapshot.setdefault("test_docs", [])
    docs_by_path = {doc["path"]: doc for doc in snapshot["test_docs"]}
    current_paths = set()
    for test in parsed.get("tests", []):
        current_paths.add(test["path"])
        doc = docs_by_path.get(test["path"])
        if not doc:
            snapshot["test_docs"].append(
                {
                    "id": new_id("testdoc"),
                    "path": test["path"],
                    "name": test["name"],
                    "command": test["command"],
                    "dataset": "",
                    "content": "",
                    "preconditions": "",
                    "result_view": "",
                    "pass_criteria": "",
                    "coverage": join_list(test.get("supported_chips", [])),
                    "owner_added": False,
                    "obsolete": False,
                }
            )
        else:
            doc["command"] = test["command"]
            doc["obsolete"] = False
    for doc in snapshot["test_docs"]:
        if not doc.get("owner_added") and doc["path"] not in current_paths:
            doc["obsolete"] = True


def update_snapshot(
    conn: sqlite3.Connection,
    release_id: str,
    app_id: str,
    mutator: Callable[[dict[str, Any]], None],
    *,
    skip_doc_deadline: bool = False,
) -> dict[str, Any]:
    with transaction(conn):
        release = get_release(conn, release_id)
        if release.get("released_locked"):
            raise RuntimeError("Release 已最终锁定，不可修改")
        if not skip_doc_deadline:
            require_can(release, "edit_snapshot", "已过 doc deadline，不可再修改文档/表单信息")
        snapshot = release["snapshots"][app_id]
        mutator(snapshot)
        save_snapshot(conn, release_id, app_id, snapshot)
        return snapshot


def sync_decision_to_later_releases(
    conn: sqlite3.Connection,
    from_release_id: str,
    app_id: str,
    decision: str,
    *,
    user: str = "system",
    role: str = "system",
) -> dict[str, Any]:
    """Copy a release_decision to every later (by created_at) release.

    A locked release is skipped. ``release`` is skipped for any release past
    its app freeze (cannot be raised back). A downgrade applies even past the
    doc deadline, matching the late-decision rule.
    """
    decision = normalize_release_decision(decision)
    result: dict[str, list] = {"applied": [], "skipped": []}
    releases = list_releases(conn)
    idx = next((i for i, r in enumerate(releases) if r["id"] == from_release_id), None)
    if idx is None:
        return result
    with transaction(conn):
        for r in releases[idx + 1:]:
            rid = r["id"]
            if r.get("released_locked"):
                result["skipped"].append({"release_id": rid, "release_name": r["name"], "reason": "已最终锁定"})
                continue
            release = get_release(conn, rid)
            snapshot = release["snapshots"].get(app_id)
            if snapshot is None:
                result["skipped"].append({"release_id": rid, "release_name": r["name"], "reason": "本 release 无此 app"})
                continue
            if decision == "release" and not can(release, "raise_to_release"):
                result["skipped"].append({"release_id": rid, "release_name": r["name"], "reason": "已过 app freeze，无法升回 release"})
                continue
            if snapshot.get("release_decision") != decision:
                snapshot["release_decision"] = decision
                snapshot["missing_items"] = missing_items_for(get_app(conn, app_id), snapshot)
                save_snapshot(conn, rid, app_id, snapshot)
            result["applied"].append({"release_id": rid, "release_name": r["name"]})
        if result["applied"]:
            audit(
                conn,
                f"同步 release 决策（{decision}）到 {len(result['applied'])} 个后续 release",
                user=user,
                role=role,
                app_id=app_id,
                release_id=from_release_id,
                event="sync_decision",
            )
    return result


def _qa_scope_additions(old_parsed: dict[str, Any], new_parsed: dict[str, Any]) -> list[str]:
    """Describe QA-scope-expanding additions: new chips or new test paths."""
    additions: list[str] = []
    old_chips = set(old_parsed.get("x86_chips", [])) | set(old_parsed.get("arm_chips", []))
    new_chips = set(new_parsed.get("x86_chips", [])) | set(new_parsed.get("arm_chips", []))
    added_chips = sorted(new_chips - old_chips)
    if added_chips:
        additions.append("新增芯片 " + ", ".join(added_chips))
    old_paths = {test.get("path") for test in old_parsed.get("tests", [])}
    new_paths = {test.get("path") for test in new_parsed.get("tests", [])}
    added_paths = sorted(path for path in new_paths - old_paths if path)
    if added_paths:
        additions.append("新增测试 " + ", ".join(added_paths))
    return additions


def apply_app_info(
    conn: sqlite3.Connection,
    release_id: str,
    app_id: str,
    raw: str | dict[str, Any],
    *,
    source: str = "upload",
    source_type: str = "owner_upload",
    commit_id: str = "",
    uploaded_by: str = "",
    role: str = "Owner",
) -> dict[str, Any]:
    release = get_release(conn, release_id)
    if release.get("released_locked"):
        raise RuntimeError("Release 已最终锁定，不可上传 app_info")
    require_can(release, "edit_app_info", "已过 doc deadline，不可再上传 app_info")
    snapshot = release["snapshots"][app_id]
    was_confirmed = bool(snapshot.get("owner_confirmed"))
    snapshot_parsed = (snapshot.get("app_info") or {}).get("parsed")
    parsed = parse_app_info(raw)
    # Owner confirmation should only become invalid when the snapshot's own
    # app_info content actually changes — a re-upload of the same file, a
    # clone from a previous release, or a fetch that returns identical
    # content must not silently force the owner to re-confirm.
    content_modified = snapshot_parsed is not None and bool(diff_app_info(snapshot_parsed, parsed))
    if snapshot.get("release_decision") == "release" and not can(release, "expand_qa_scope"):
        current_parsed = (snapshot.get("app_info") or {}).get("parsed")
        if current_parsed is not None:
            additions = _qa_scope_additions(current_parsed, parsed)
            if additions:
                raise RuntimeError(
                    "已过 app 冻结 deadline，新 app_info 会扩大 QA 范围（"
                    + "；".join(additions)
                    + "）。如确需新增，请联系 RM 调整 app 冻结 deadline。"
                )
    previous = previous_release(conn, release_id)
    old_parsed = None
    if previous and app_id in previous["snapshots"]:
        old_parsed = (previous["snapshots"][app_id].get("app_info") or {}).get("parsed")
    if old_parsed is None:
        old_parsed = (snapshot.get("app_info") or {}).get("parsed")
    diffs = diff_app_info(old_parsed, parsed) if old_parsed is not None else []
    snapshot["app_info"] = {
        "source": source,
        "source_type": source_type,
        "synced_at": now(),
        "commit_id": commit_id,
        "uploaded_by": uploaded_by,
        "raw": parsed["raw"],
        "parsed": parsed,
    }
    snapshot["app_info_diffs"] = diffs
    snapshot["version"] = parsed.get("app_version") or snapshot.get("version", "")
    snapshot["x86_chips"] = ",".join(order_chips(parsed.get("x86_chips", [])))
    snapshot["arm_chips"] = ",".join(order_chips(parsed.get("arm_chips", [])))
    snapshot["python_labels"] = ",".join(parsed.get("python_labels", []))
    snapshot["pytorch_labels"] = ",".join(parsed.get("pytorch_labels", []))
    snapshot["build_os"] = ",".join(parsed.get("build_os", []))
    snapshot["build_arches"] = ",".join(parsed.get("build_arches", []))
    ensure_test_docs(snapshot, parsed, diffs)
    if was_confirmed and content_modified:
        snapshot["owner_confirmed"] = False
    # build_targets / test_targets are coarse list-of-dict aggregates; the
    # readable per-field diffs (version, chips, test_cmd*) cover the same ground.
    detail = [
        {
            "field": d.get("field", ""),
            "label": f"{d.get('type', '')}（{d.get('field', '')}）",
            "old": fmt_audit_value(d.get("old_value")),
            "new": fmt_audit_value(d.get("new_value")),
        }
        for d in diffs
        if d.get("field") not in ("build_targets", "test_targets")
    ]
    with transaction(conn):
        save_snapshot(conn, release_id, app_id, snapshot)
        audit(
            conn,
            f"{app_id} 更新 app_info.json，差异 {len(diffs)} 项",
            user=uploaded_by or "system",
            role=role,
            app_id=app_id,
            release_id=release_id,
            event="upload_app_info",
            detail=detail,
        )
        if was_confirmed and content_modified:
            audit(
                conn,
                f"{app_id} Owner 确认因 app_info 更新自动失效",
                user=uploaded_by or "system",
                role=role,
                app_id=app_id,
                release_id=release_id,
                event="owner_confirm_invalidated",
                detail=[{"field": "owner_confirmed", "label": "Owner 确认",
                         "old": "已确认", "new": "未确认（app_info 更新自动失效）"}],
            )
    return snapshot


def missing_item_text(item: Any) -> str:
    """Display text for a missing_items entry.

    Items are stored as ``{"kind": "doc"|"qa", "text": str}``. Older snapshots
    or legacy callers may still pass bare strings; both are handled here so
    downstream display / equality checks keep working through any rolling
    migration.
    """
    if isinstance(item, dict):
        return str(item.get("text", ""))
    return str(item)


def missing_item_kind(item: Any) -> str:
    """Return the kind of a missing_items entry (``"doc"`` or ``"qa"``).

    Falls back to inspecting the text prefix when given a legacy string item,
    so an old snapshot that wasn't refreshed yet still gates correctly.
    """
    if isinstance(item, dict):
        return str(item.get("kind", "doc"))
    return "qa" if str(item).startswith("QA ") else "doc"


def missing_items_for(app: dict[str, Any], snapshot: dict[str, Any]) -> list[dict[str, str]]:
    """Readiness and final-release gate items shown to RM/owners.

    Each entry is ``{"kind": "doc"|"qa", "text": str}``. ``doc`` entries
    block ``_qualifies_for_final``; ``qa`` entries are informational and do
    not (QA status itself is the gate). Splitting by kind avoids the
    string-prefix coupling that broke when an owner named a test doc
    starting with "QA ".
    """
    decision = normalize_release_decision(snapshot.get("release_decision"))
    if decision != "release":
        return []
    missing: list[dict[str, str]] = []

    def add_doc(text: str) -> None:
        missing.append({"kind": "doc", "text": text})

    def add_qa(text: str) -> None:
        missing.append({"kind": "qa", "text": text})

    if not snapshot.get("owners"):
        add_doc("缺少 owner")
    if not app.get("git_url"):
        add_doc("缺少 Gerrit URL")
    if not app.get("git_branch"):
        add_doc("缺少 branch")
    if not (snapshot.get("official_name") or "").strip():
        add_doc("缺少官方名称")
    if not (snapshot.get("type") or "").strip():
        add_doc("缺少 App类型")
    description = (snapshot.get("description") or "").strip()
    if not description:
        add_doc("缺少描述（30字内）")
    elif app_description_count(description) > MAX_APP_DESCRIPTION_CHARS:
        add_doc("描述超过30字")
    if not snapshot.get("app_info"):
        add_doc("缺少可追溯 AppInfoSnapshot")
    if not snapshot.get("version"):
        add_doc("缺少 对应官方版本")
    if not snapshot.get("x86_chips"):
        add_doc("缺少 X86支持芯片系列")
    if normalize_doc_target(snapshot.get("doc_target")) in DOC_TARGETS:
        doc = snapshot.get("doc", {})
        required = {
            "intro": "基本介绍",
            "image_usage": "镜像使用方法",
            "binary_usage": "二进制包使用方法",
            "env_setup": "环境搭建",
        }
        for key, label in required.items():
            if not doc.get(key):
                add_doc(f"缺少{label}")
    for doc in snapshot.get("test_docs", []):
        if doc.get("obsolete"):
            continue
        if doc.get("owner_added") and not doc.get("command"):
            add_doc(f"{doc['path']} 缺少 owner-added 测试命令")
        for key, label in {"dataset": "测试数据集", "content": "测试内容", "result_view": "结果查看方式", "pass_criteria": "通过标准"}.items():
            if not doc.get(key):
                add_doc(f"{doc['path']} 缺少{label}")
    if not snapshot.get("owner_confirmed"):
        add_doc("Owner 未确认 doc")
    qa_status = snapshot.get("qa_status", "not_checked")
    if qa_status == "not_checked":
        add_qa("QA 未测试")
    elif qa_status == "cannot_release":
        add_qa("QA 标注为不可发布")
    return missing


def refresh_missing_items(conn: sqlite3.Connection, release_id: str) -> dict[str, list[str]]:
    """Recompute missing_items for every snapshot in the release; return map.

    Only writes back when the recomputed value (or normalized decision) differs
    from what is stored, so /api/state polling doesn't thrash the DB.
    """
    release = get_release(conn, release_id)
    if release.get("released_locked"):
        return {app_id: snap.get("missing_items", []) for app_id, snap in release["snapshots"].items()}
    apps = {app["id"]: app for app in list_apps(conn)}
    results: dict[str, list[str]] = {}
    with transaction(conn):
        for app_id, snapshot in release["snapshots"].items():
            app = apps.get(app_id)
            if not app:
                continue
            before = dumps_json(snapshot)
            snapshot["release_decision"] = normalize_release_decision(snapshot.get("release_decision"))
            snapshot.pop("cicd", None)
            items = missing_items_for(app, snapshot)
            snapshot["missing_items"] = items
            results[app_id] = items
            after = dumps_json(snapshot)
            if before != after:
                save_snapshot(conn, release_id, app_id, snapshot)
    return results


def qa_set_status(
    conn: sqlite3.Connection,
    release_id: str,
    app_id: str,
    status: str,
    *,
    issue_note: str = "",
    user: str = "qa",
    role: str = "QA",
) -> dict[str, Any]:
    if status not in QA_STATUSES:
        raise ValueError(f"Invalid QA status: {status}")
    release = get_release(conn, release_id)
    if release.get("released_locked"):
        raise RuntimeError("Release 已最终锁定，不可修改 QA 状态")
    snapshot = release["snapshots"].get(app_id)
    if not snapshot:
        raise KeyError(f"App {app_id} not in release")
    if snapshot.get("release_decision") != "release":
        raise RuntimeError("仅 release 决策的 app 可由 QA 标注状态")
    if status == "has_issues" and not (issue_note or "").strip():
        raise ValueError("标注「存在问题」时必须填写问题说明")
    old_status = snapshot.get("qa_status", "not_checked")
    old_note = snapshot.get("qa_issue_note", "")
    snapshot["qa_status"] = status
    snapshot["qa_issue_note"] = (issue_note or "").strip() if status == "has_issues" else ""
    detail = [{"field": "qa_status", "label": "QA 状态", "old": old_status, "new": status}]
    if old_note or snapshot["qa_issue_note"]:
        detail.append({"field": "qa_issue_note", "label": "问题说明", "old": old_note, "new": snapshot["qa_issue_note"]})
    with transaction(conn):
        save_snapshot(conn, release_id, app_id, snapshot)
        audit(
            conn,
            f"QA 标注 {app_id} 为 {status}" + (f"：{issue_note}" if issue_note else ""),
            user=user,
            role=role,
            app_id=app_id,
            release_id=release_id,
            event="qa_set_status",
            detail=detail,
        )
    return snapshot


QA_TEST_RESULT_STATUSES = {"pass", "fail", "skip", "unknown"}


def _normalize_test_results(raw: Any) -> list[dict[str, Any]]:
    """Coerce a batch item's test_results into a clean list-of-dicts shape.

    Unknown fields are dropped so the LLM cannot smuggle arbitrary keys into
    the snapshot; status is clamped to QA_TEST_RESULT_STATUSES.
    """
    if not isinstance(raw, list):
        return []
    cleaned: list[dict[str, Any]] = []
    for row in raw:
        if not isinstance(row, dict):
            continue
        status = str(row.get("status") or "unknown").lower()
        if status not in QA_TEST_RESULT_STATUSES:
            status = "unknown"
        cleaned.append({
            "test": str(row.get("test") or "").strip(),
            "arch": str(row.get("arch") or "").strip(),
            "status": status,
            "perf": str(row.get("perf") or "").strip(),
            "note": str(row.get("note") or "").strip(),
        })
    return cleaned


def qa_set_status_batch(
    conn: sqlite3.Connection,
    release_id: str,
    items: list[dict[str, Any]],
    *,
    user: str = "qa",
    role: str = "QA",
) -> dict[str, dict[str, Any]]:
    """Apply several QA-status updates atomically.

    Every item is validated first; if any item is invalid the whole batch is
    rejected and nothing is written. Only on full success is a single commit
    issued, so a mid-batch failure can never leave a partial save.
    """
    release = get_release(conn, release_id)
    if release.get("released_locked"):
        raise RuntimeError("Release 已最终锁定，不可修改 QA 状态")
    # (app_id, snapshot, status, issue_note, old_status, old_note)
    prepared: list[tuple[str, dict[str, Any], str, str, str, str]] = []
    errors: list[str] = []
    for item in items:
        app_id = item.get("app_id", "")
        status = item.get("status", "")
        issue_note = (item.get("issue_note") or "").strip()
        if status not in QA_STATUSES:
            errors.append(f"{app_id}：无效的 QA 状态 {status!r}")
            continue
        snapshot = release["snapshots"].get(app_id)
        if not snapshot:
            errors.append(f"{app_id}：不在本 release 中")
            continue
        if snapshot.get("release_decision") != "release":
            errors.append(f"{app_id}：仅 release 决策的 app 可标注 QA 状态")
            continue
        if status == "has_issues" and not issue_note:
            errors.append(f"{app_id}：标注「存在问题」时必须填写问题说明")
            continue
        prepared.append((app_id, snapshot, status, issue_note, snapshot.get("qa_status", "not_checked"), snapshot.get("qa_issue_note", "")))
    if errors:
        raise ValueError("；".join(errors))
    with transaction(conn):
        for app_id, snapshot, status, issue_note, old_status, old_note in prepared:
            snapshot["qa_status"] = status
            snapshot["qa_issue_note"] = issue_note if status == "has_issues" else ""
            save_snapshot(conn, release_id, app_id, snapshot)
            detail = [{"field": "qa_status", "label": "QA 状态", "old": old_status, "new": status}]
            if old_note or snapshot["qa_issue_note"]:
                detail.append({"field": "qa_issue_note", "label": "问题说明", "old": old_note, "new": snapshot["qa_issue_note"]})
            audit(
                conn,
                f"QA 标注 {app_id} 为 {status}" + (f"：{issue_note}" if issue_note else ""),
                user=user,
                role=role,
                app_id=app_id,
                release_id=release_id,
                event="qa_set_status",
                detail=detail,
            )
    return {app_id: snapshot for app_id, snapshot, *_ in prepared}


def qa_log_dir(db_path: str | Path) -> Path:
    base = Path(db_path).resolve().parent / "qa_logs"
    base.mkdir(exist_ok=True)
    return base


def qa_upload_log(
    conn: sqlite3.Connection,
    db_path: str | Path,
    release_id: str,
    content: bytes,
    filename: str,
    *,
    user: str = "qa",
    role: str = "QA",
) -> dict[str, str]:
    release = get_release(conn, release_id)
    if release.get("released_locked"):
        raise RuntimeError("Release 已最终锁定，不可上传 QA log")
    if not filename:
        raise ValueError("filename required")
    safe_name = re.sub(r"[^A-Za-z0-9._-]+", "_", filename) or "qa_log"
    target_dir = qa_log_dir(db_path)
    storage_path = target_dir / f"{release_id}__{safe_name}"
    storage_path.write_bytes(content)
    with transaction(conn):
        conn.execute(
            """
            INSERT INTO qa_logs(release_id, filename, storage_path, uploaded_at, uploaded_by)
            VALUES (?, ?, ?, ?, ?)
            ON CONFLICT(release_id) DO UPDATE SET
              filename=excluded.filename,
              storage_path=excluded.storage_path,
              uploaded_at=excluded.uploaded_at,
              uploaded_by=excluded.uploaded_by
            """,
            (release_id, safe_name, str(storage_path), now(), user),
        )
        audit(
            conn,
            f"QA 上传 log：{safe_name}",
            user=user,
            role=role,
            release_id=release_id,
            event="qa_upload_log",
        )
    return {"filename": safe_name, "storage_path": str(storage_path), "uploaded_at": now(), "uploaded_by": user}


def get_qa_log(conn: sqlite3.Connection, release_id: str) -> dict[str, str] | None:
    row = conn.execute(
        "SELECT release_id, filename, storage_path, uploaded_at, uploaded_by FROM qa_logs WHERE release_id = ?",
        (release_id,),
    ).fetchone()
    return dict(row) if row else None


def _qa_analysis_inventory(release: dict[str, Any], apps: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    """Build the per-app test inventory the LLM uses as its target schema.

    Only release-decision=release apps are included; each app lists its
    enabled, non-weekly tests with the test_cmd string so the model can match
    log lines back to a specific (app, test, arch).
    """
    inventory: list[dict[str, Any]] = []
    for app_id, snapshot in release["snapshots"].items():
        if snapshot.get("release_decision") != "release":
            continue
        app = apps.get(app_id) or {}
        tests = []
        for test in snapshot.get("tests") or []:
            if not test.get("enabled", True):
                continue
            tests.append({
                "name": test.get("name") or test.get("path") or "",
                "path": test.get("path") or "",
                "command": test.get("command") or "",
                "supported_chips": test.get("supported_chips") or [],
                "arches": test.get("arch_list") or [],
            })
        inventory.append({
            "app_id": app_id,
            "app_name": snapshot.get("app_name") or app.get("name") or app_id,
            "version": snapshot.get("version") or snapshot.get("app_version") or "",
            "tests": tests,
        })
    return inventory


_QA_ANALYSIS_SYSTEM = (
    "You analyze QA test logs/summary tables for an HPC release. Match the "
    "log content to the provided per-app test inventory and report results.\n"
    "Return STRICT JSON of the form: "
    "{\"apps\": [{\"app_id\": str, "
    "\"qa_status\": \"qa_passed\"|\"has_issues\"|\"cannot_release\"|\"not_checked\", "
    "\"qa_issue_note\": str, \"tests\": [{\"test\": str, \"arch\": str, "
    "\"status\": \"pass\"|\"fail\"|\"skip\"|\"unknown\", \"perf\": str, \"note\": str}]}]}.\n"
    "Rules:\n"
    "- Include EVERY app_id from the inventory. If the log has no data for an app, set its tests to status=unknown and qa_status=not_checked.\n"
    "- qa_status=qa_passed only when every listed test for that app shows a clear pass in the log.\n"
    "- qa_status=has_issues when any test fails or shows clear regression; qa_issue_note must concisely list which tests/arches failed and why (in Chinese).\n"
    "- qa_status=cannot_release only when the log explicitly says release is blocked.\n"
    "- Match app/test names case-insensitively and tolerate small spelling variants. The log may be plain text OR a multi-sheet spreadsheet rendered as TSV (each sheet preceded by `### Sheet: <name> ###`).\n"
    "- perf: short numeric/throughput summary if the log contains one, else empty.\n"
    "- note: at most one short Chinese sentence per test (cause of fail, perf delta, or empty).\n"
    "- Do NOT invent tests not in the inventory. Output JSON only, no prose, no code fences."
)


def _xlsx_to_text(raw: bytes, *, max_rows_per_sheet: int = 2000) -> str:
    """Render an xlsx workbook as TSV-ish plain text the LLM can read.

    Each sheet is preceded by `### Sheet: <name> ###`; rows are tab-joined and
    trailing empty cells are stripped so the prompt stays compact. Fails fast
    with a clear message if the file isn't a valid xlsx.
    """
    import io as _io
    from openpyxl import load_workbook

    wb = load_workbook(_io.BytesIO(raw), data_only=True, read_only=True)
    chunks: list[str] = []
    for name in wb.sheetnames:
        ws = wb[name]
        chunks.append(f"### Sheet: {name} ###")
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_rows_per_sheet:
                chunks.append(f"...[truncated after {max_rows_per_sheet} rows]...")
                break
            cells: list[str] = []
            for v in row:
                if v is None:
                    cells.append("")
                else:
                    # Cells legitimately contain tabs/newlines (multi-line perf
                    # notes etc.), so neutralize them to keep the TSV unambiguous.
                    cells.append(str(v).replace("\t", " ").replace("\r", " ").replace("\n", " "))
            while cells and cells[-1] == "":
                cells.pop()
            if cells:
                chunks.append("\t".join(cells))
        chunks.append("")
    wb.close()
    return "\n".join(chunks)


def _qa_log_to_text(path: Path, raw: bytes) -> str:
    """Decode a QA-log file for the LLM. xlsx → TSV; everything else → utf-8."""
    if path.suffix.lower() == ".xlsx":
        try:
            return _xlsx_to_text(raw)
        except Exception as exc:
            raise RuntimeError(f"无法解析 xlsx：{exc}") from exc
    return raw.decode("utf-8", errors="replace")


def _qa_progress(progress: Callable[..., None] | None, stage: str, message: str, **extra: Any) -> None:
    if progress:
        if extra:
            try:
                progress(stage, message, **extra)
                return
            except TypeError:
                pass
        progress(stage, message)


def qa_analyze_log(
    conn: sqlite3.Connection,
    db_path: str | Path,
    release_id: str,
    *,
    llm_call: Callable[[str, str], str] | None = None,
    max_log_chars: int = 200_000,
    progress: Callable[[str, str], None] | None = None,
    max_llm_attempts: int = 2,
) -> dict[str, Any]:
    """Ask the LLM to summarize the uploaded QA log against the test inventory.

    Pure read: no DB writes, no snapshot mutation. The caller (server handler)
    returns the result to the UI which prefills the QA edit form, and only the
    QA-confirmed values are later persisted via qa_set_status_batch.
    """
    _qa_progress(progress, "checking_log", "正在检查已上传的 QA log")
    meta = get_qa_log(conn, release_id)
    if not meta:
        raise RuntimeError("本 release 还未上传 QA log")
    log_path = Path(meta["storage_path"])
    if not log_path.exists():
        raise RuntimeError("QA log 文件丢失")
    _qa_progress(progress, "reading_log", f"正在读取 QA log：{meta.get('filename', log_path.name)}")
    raw = log_path.read_bytes()
    if log_path.suffix.lower() == ".xlsx":
        _qa_progress(progress, "parsing_excel", "正在解析 Excel 文件")
    else:
        _qa_progress(progress, "parsing_text", "正在解析文本 log")
    text = _qa_log_to_text(log_path, raw)
    truncated = False
    if len(text) > max_log_chars:
        _qa_progress(progress, "truncating_log", "log 较大，正在截断中段以控制 LLM 上下文")
        head = text[: max_log_chars // 2]
        tail = text[-max_log_chars // 2 :]
        text = head + "\n\n...[log truncated]...\n\n" + tail
        truncated = True

    _qa_progress(progress, "building_inventory", "正在准备 app 和测试清单")
    release = get_release(conn, release_id)
    apps = {app["id"]: app for app in list_apps(conn)}
    inventory = _qa_analysis_inventory(release, apps)
    if not inventory:
        raise RuntimeError("本 release 没有 release 决策为 release 的 app，无法分析")

    _qa_progress(progress, "building_prompt", "正在构造 LLM 分析上下文")
    user_payload = json.dumps({"inventory": inventory, "log": text}, ensure_ascii=False)

    if llm_call is None:
        from release_system.llm import chat_json

        def default_llm_call(system: str, payload: str) -> str:
            def stream_progress(token_count: int) -> None:
                _qa_progress(
                    progress,
                    "streaming_llm",
                    f"正在接收 LLM 输出：已收到 {token_count} token",
                    token_count=token_count,
                )

            return chat_json(system, payload, progress=stream_progress)

        llm_call = default_llm_call

    max_llm_attempts = max(1, max_llm_attempts)
    parsed: dict[str, Any] | None = None
    for attempt in range(1, max_llm_attempts + 1):
        try:
            suffix = f"（第 {attempt}/{max_llm_attempts} 次）" if max_llm_attempts > 1 else ""
            _qa_progress(progress, "waiting_llm", f"正在等待 LLM 返回结果{suffix}", token_count=0)
            raw_reply = llm_call(_QA_ANALYSIS_SYSTEM, user_payload)
            _qa_progress(progress, "parsing_llm", "正在解析 LLM 返回结果")
            try:
                parsed = json.loads(raw_reply)
            except json.JSONDecodeError:
                # Some local models wrap JSON in ```json fences — strip them and retry.
                cleaned = re.sub(r"^```(?:json)?\s*|\s*```$", "", raw_reply.strip(), flags=re.MULTILINE)
                parsed = json.loads(cleaned)
            break
        except Exception as exc:
            if attempt >= max_llm_attempts:
                raise
            _qa_progress(progress, "retrying_llm", f"LLM 调用失败，正在重试：{exc}")
            time.sleep(min(attempt, 3))

    if parsed is None:
        raise RuntimeError("LLM 未返回有效结果")

    _qa_progress(progress, "normalizing_result", "正在整理 AI 建议")
    valid_ids = {entry["app_id"] for entry in inventory}
    apps_out: list[dict[str, Any]] = []
    for row in parsed.get("apps") or []:
        if not isinstance(row, dict):
            continue
        app_id = row.get("app_id")
        if app_id not in valid_ids:
            continue
        status = row.get("qa_status") or "not_checked"
        if status not in QA_STATUSES:
            status = "not_checked"
        apps_out.append({
            "app_id": app_id,
            "qa_status": status,
            "qa_issue_note": str(row.get("qa_issue_note") or "").strip(),
            "test_results": _normalize_test_results(row.get("tests")),
        })

    with transaction(conn):
        audit(
            conn,
            f"QA AI 分析 log：{meta.get('filename', '')}",
            user="qa-ai",
            role="QA",
            release_id=release_id,
            event="qa_analyze_log",
        )
    _qa_progress(progress, "completed", "AI 分析完成")
    return {"apps": apps_out, "log_truncated": truncated, "log_chars": len(raw)}


def export_test_scope_csv(conn: sqlite3.Connection, release_id: str) -> str:
    """Return CSV text of release-decision=release apps in the release."""
    release = get_release(conn, release_id)
    apps = {app["id"]: app for app in list_apps(conn)}
    out = io.StringIO()
    writer = csv.writer(out)
    writer.writerow(["app_name", "version", "gerrit_url", "branch", "owners"])
    rows = []
    for app_id, snapshot in release["snapshots"].items():
        if snapshot.get("release_decision") != "release":
            continue
        app = apps.get(app_id)
        if not app:
            continue
        view = app_view(app, snapshot)
        rows.append(
            (
                view["name"],
                snapshot.get("version", ""),
                view["git_url"],
                view["git_branch"],
                ",".join(view["owners"]),
            )
        )
    rows.sort(key=lambda r: r[0].lower())
    for row in rows:
        writer.writerow(row)
    return out.getvalue()


# --- QA reports: release report + test command ------------------------------
# These mirror the offline get_release_report_test_cmd.py script, but draw on
# the app_info already uploaded into each release snapshot instead of fetching
# app_info.json from Gerrit.

QA_RELEASE_REPORT_COLUMNS = [
    "类别", "名称", "Owner", "类型", "描述", "官方URL", "git_url", "git_branch",
    "对应官方版本", "X86支持芯片系列", "ARM支持芯片类型", "对比",
    "开发者社区发布情况", "开发者社区发布包支持python版本",
    "开发者社区发布包支持的底层框架及版本",
    "ARM / Kylin sanity", "Ubuntu / 兼容性 sanity",
]

QA_TEST_CMD_COLUMNS = [
    "app_name", "git_branch", "app_version", "arch",
    "maca_version", "test_name", "docker_cmd",
]


def _report_normalize_arch(value: Any) -> str:
    """Normalize arch aliases (arm/x86) to canonical arm64/amd64."""
    a = str(value or "").strip().lower()
    if a in ("arm", "arm64", "aarch64"):
        return "arm64"
    if a in ("x86", "x86_64", "amd64"):
        return "amd64"
    return a


def _report_denormalize_arch(value: Any) -> str:
    """Map canonical arm64/amd64 back to the short arm/x86 used in the CSV."""
    n = _report_normalize_arch(value)
    if n == "arm64":
        return "arm"
    if n == "amd64":
        return "x86"
    return n


def _report_build_arches(app_info: dict[str, Any]) -> set[str]:
    """Normalized arches declared in app_info.app_build.*.arch."""
    arches: set[str] = set()
    for cfg in (app_info.get("app_build") or {}).values():
        if not isinstance(cfg, dict):
            continue
        a = _report_normalize_arch(cfg.get("arch"))
        if a:
            arches.add(a)
    return arches


def _report_test_arches(test_cfg: dict[str, Any]) -> set[str]:
    """Arches a test supports, taken from the build_key suffix in
    supported_chip: {chip: [build_key, ...]} (e.g. ubuntu20.04_amd64)."""
    arches: set[str] = set()
    sc = test_cfg.get("supported_chip")
    if isinstance(sc, dict):
        for build_keys in sc.values():
            if not isinstance(build_keys, list):
                continue
            for key in build_keys:
                a = _report_normalize_arch(str(key).rsplit("_", 1)[-1])
                if a:
                    arches.add(a)
    return arches


def _report_test_skip(test_cfg: dict[str, Any], arch: str) -> bool:
    """True when supported_chip is a non-empty dict that excludes this arch
    on every chip — the test explicitly does not support the arch."""
    sc = test_cfg.get("supported_chip")
    target = _report_normalize_arch(arch)
    if not (isinstance(sc, dict) and sc) or not target:
        return False
    for build_keys in sc.values():
        keys = build_keys if isinstance(build_keys, list) else []
        if any(_report_normalize_arch(str(k).rsplit("_", 1)[-1]) == target for k in keys):
            return False
    return True


def _report_docker_cmd(test_cfg: dict[str, Any]) -> str:
    """Assemble the `docker run ...` command for one app_test entry."""
    container_args = str(test_cfg.get("container_args") or "").strip()
    test_cmd = str(test_cfg.get("test_cmd") or "").strip()
    img_target = str(test_cfg.get("img_target") or "").strip().lower()
    image = f"[docker_image_{img_target}]" if img_target else "[docker_image]"
    parts = ["docker run --pull always --rm -e MACA_PERF_DIR=/tmp"]
    if test_cfg.get("mount_dataset"):
        parts.append("-v /pde_hpc/dataset:/hpc_dataset:ro")
    if container_args:
        parts.append(container_args)
    parts.append(image)
    parts.append(f"sh -c '{test_cmd}'")
    return " ".join(parts)


def _report_test_cmd_rows(
    raw: dict[str, Any], app_name: str, git_branch: str, maca_version: str
) -> list[list[str]]:
    """Build test-command rows for one app from its raw app_info.json dict.

    Only enabled, non-weekly tests are kept; each (test, arch) is one row.
    """
    version_value = raw.get("app_version")
    app_version = str(version_value).strip() if version_value not in (None, "") else ""
    app_arches = _report_build_arches(raw)

    rows: list[list[str]] = []
    for test_name, test_cfg in (raw.get("app_test") or {}).items():
        if not isinstance(test_cfg, dict) or not test_cfg.get("enabled"):
            continue
        if str(test_cfg.get("test_period") or "").strip().lower() == "weekly":
            continue
        if test_cfg.get("ignore_release"):
            continue
        docker_cmd = _report_docker_cmd(test_cfg)
        for n_arch in sorted(a for a in (_report_test_arches(test_cfg) or app_arches) if a):
            if _report_test_skip(test_cfg, n_arch):
                continue
            rows.append([
                app_name,
                git_branch,
                app_version,
                _report_denormalize_arch(n_arch),
                maca_version,
                str(test_name),
                docker_cmd,
            ])
    return rows


def _compare_summary(
    snapshot: dict[str, Any], base_snapshot: dict[str, Any] | None
) -> str:
    """Summarize what changed for one app between two releases.

    Output is a short, human-readable string like
    '新增发布; 支持芯片修改; 测试命令改变' for the 对比 column.
    """
    cur_decision = normalize_release_decision(snapshot.get("release_decision"))
    if base_snapshot is None:
        return "新增发布" if cur_decision == "release" else ""

    tags: list[str] = []
    base_decision = normalize_release_decision(base_snapshot.get("release_decision"))
    if cur_decision == "release" and base_decision != "release":
        tags.append("新增发布")
    elif cur_decision != "release" and base_decision == "release":
        tags.append("停止发布")

    def _chip_set(value: Any) -> tuple[str, ...]:
        return tuple(order_chips(value or ""))

    if (_chip_set(snapshot.get("x86_chips")) != _chip_set(base_snapshot.get("x86_chips"))
            or _chip_set(snapshot.get("arm_chips")) != _chip_set(base_snapshot.get("arm_chips"))):
        tags.append("支持芯片修改")

    def _test_cmd_set(snap: dict[str, Any]) -> set[tuple[str, str]]:
        raw = (snap.get("app_info") or {}).get("raw")
        if not isinstance(raw, dict):
            return set()
        out: set[tuple[str, str]] = set()
        for name, cfg in (raw.get("app_test") or {}).items():
            if not isinstance(cfg, dict) or not cfg.get("enabled"):
                continue
            if str(cfg.get("test_period") or "").strip().lower() == "weekly":
                continue
            if cfg.get("ignore_release"):
                continue
            out.add((str(name), _report_docker_cmd(cfg)))
        return out

    if _test_cmd_set(snapshot) != _test_cmd_set(base_snapshot):
        tags.append("测试范围变更")

    def _test_doc_map(snap: dict[str, Any]) -> dict[str, tuple[str, str, str, str, str]]:
        out: dict[str, tuple[str, str, str, str, str]] = {}
        for d in snap.get("test_docs") or []:
            if not isinstance(d, dict) or d.get("obsolete"):
                continue
            path = str(d.get("path") or "")
            if not path:
                continue
            out[path] = (
                str(d.get("dataset") or ""),
                str(d.get("content") or ""),
                str(d.get("result_view") or ""),
                str(d.get("pass_criteria") or ""),
                str(d.get("command") or ""),
            )
        return out

    if _test_doc_map(snapshot) != _test_doc_map(base_snapshot):
        tags.append("测试说明变更")

    if (snapshot.get("version") or "") != (base_snapshot.get("version") or ""):
        tags.append("版本变更")

    return "; ".join(tags)


def build_qa_reports(
    conn: sqlite3.Connection,
    release_id: str,
    compare_release_id: str | None = None,
) -> dict[str, Any]:
    """Build the release report + test command tables for a release.

    release_report: one catalog-style row per app in the release (column
    layout follows release_report.csv; 类别 is not tracked and stays
    blank; 对比 is a short summary against compare_release_id when given,
    the rest — incl. 社区发布信息 and sanity — comes from the snapshot).
    test_cmd: one row per (test, arch) drawn from each app's uploaded
    app_info, matching get_release_report_test_cmd.py's test command output.
    """
    release = get_release(conn, release_id)
    apps = {app["id"]: app for app in list_apps(conn)}
    maca_version = release.get("maca_version", "")

    base_snapshots: dict[str, dict[str, Any]] = {}
    base_release_name = ""
    if compare_release_id and compare_release_id != release_id:
        try:
            base_release = get_release(conn, compare_release_id)
            base_snapshots = base_release.get("snapshots") or {}
            base_release_name = base_release.get("name", "")
        except Exception:
            base_snapshots = {}
            base_release_name = ""

    items = []
    for app_id, snapshot in release["snapshots"].items():
        app = apps.get(app_id)
        if app:
            items.append((app_view(app, snapshot), app, snapshot, app_id))
    # release apps first, then non-release (cicd_only / stopped) at the bottom;
    # within each bucket, alphabetical by name.
    items.sort(key=lambda t: (
        0 if normalize_release_decision(t[2].get("release_decision")) == "release" else 1,
        (t[0]["name"] or "").lower(),
    ))

    release_rows: list[list[str]] = []
    release_rows_meta: list[dict[str, Any]] = []
    test_rows: list[list[str]] = []
    compare_active = bool(compare_release_id) and compare_release_id != release_id
    for view, app, snapshot, app_id in items:
        community = snapshot.get("community") or {}
        sanity = snapshot.get("sanity") or {}
        compare_value = _compare_summary(snapshot, base_snapshots.get(app_id)) if compare_active else ""
        decision = normalize_release_decision(snapshot.get("release_decision"))
        release_rows_meta.append({"release_decision": decision, "is_release": decision == "release"})
        release_rows.append([
            "AI4Sci" if view["doc_target"] == "ai4sci" else "HPC",  # 类别
            view["official_name"],
            ",".join(view["owners"]),
            view["type"],
            view["description"],
            view["official_url"],
            app.get("git_url", ""),
            app.get("git_branch", ""),
            snapshot.get("version", ""),
            ",".join(order_chips(snapshot.get("x86_chips", ""))),
            ",".join(order_chips(snapshot.get("arm_chips", ""))),
            compare_value,
            community.get("release_status", ""),
            community.get("python_version", ""),
            community.get("framework_version", ""),
            "✔" if sanity.get("arm_kylin") else "",
            "✔" if sanity.get("ubuntu") else "",
        ])
        if decision == "release":
            raw = (snapshot.get("app_info") or {}).get("raw")
            if isinstance(raw, dict):
                test_rows.extend(
                    _report_test_cmd_rows(
                        raw, view["official_name"], app.get("git_branch", ""), maca_version
                    )
                )
    test_rows.sort(key=lambda r: (r[0].lower(), r[1].lower(), r[2].lower(), r[3].lower()))

    return {
        "release_name": release.get("name", ""),
        "maca_version": maca_version,
        "compare_release_id": compare_release_id or "",
        "compare_release_name": base_release_name,
        "release_report": {
            "columns": QA_RELEASE_REPORT_COLUMNS,
            "rows": release_rows,
            "rows_meta": release_rows_meta,
        },
        "test_cmd": {"columns": QA_TEST_CMD_COLUMNS, "rows": test_rows},
    }


def final_lock_release(conn: sqlite3.Connection, release_id: str, *, user: str = "rm", role: str = "RM") -> dict[str, str]:
    """Final lock: freeze all writes, generate final artifacts.

    App metadata is already per-release in the snapshot, so the snapshot is
    self-contained once locked -- no separate app_meta copy is needed.
    """
    release = get_release(conn, release_id)
    if release.get("released_locked"):
        raise RuntimeError("Release 已最终锁定")
    with transaction(conn):
        refresh_missing_items(conn, release_id)
        release = get_release(conn, release_id)
        for app_id, snapshot in release["snapshots"].items():
            if _qualifies_for_final(snapshot):
                snapshot["locked_in_release"] = True
            save_snapshot_raw(conn, release_id, app_id, snapshot)
        conn.execute(
            "UPDATE releases SET released_locked = 1, released_locked_at = ?, released_locked_by = ? WHERE id = ?",
            (now(), user, release_id),
        )
        artifacts = generate_artifacts(conn, release_id, final=True, from_lock=True)
        audit(conn, f"Release 最终锁定：{release['name']}", user=user, role=role, release_id=release_id, event="final_lock")
    return artifacts


def final_unlock_release(conn: sqlite3.Connection, release_id: str, *, user: str = "rm", role: str = "RM") -> None:
    """Reverse a final lock: clear the locked flag, delete final artifacts."""
    release = get_release(conn, release_id)
    if not release.get("released_locked"):
        raise RuntimeError("Release 未锁定，无需解锁")
    with transaction(conn):
        conn.execute(
            "UPDATE releases SET released_locked = 0, released_locked_at = '', released_locked_by = '' WHERE id = ?",
            (release_id,),
        )
        for app_id, snapshot in release["snapshots"].items():
            snapshot.pop("app_meta", None)
            snapshot.pop("locked_in_release", None)
            save_snapshot_raw(conn, release_id, app_id, snapshot)
        conn.execute("DELETE FROM artifacts WHERE release_id = ? AND final = 1", (release_id,))
        audit(conn, f"Release 解锁：{release['name']}", user=user, role=role, release_id=release_id, event="final_unlock")


def save_snapshot_raw(conn: sqlite3.Connection, release_id: str, app_id: str, snapshot: dict[str, Any]) -> None:
    """Save snapshot WITHOUT lock checks; used by lock/unlock themselves."""
    conn.execute(
        """
        INSERT INTO snapshots(release_id, app_id, data_json)
        VALUES (?, ?, ?)
        ON CONFLICT(release_id, app_id) DO UPDATE SET data_json=excluded.data_json
        """,
        (release_id, app_id, dumps_json(snapshot)),
    )


def _qualifies_for_final(snapshot: dict[str, Any]) -> bool:
    """True if this snapshot should be included in the final release_note/manuals."""
    if snapshot.get("release_decision") != "release":
        return False
    if not snapshot.get("owner_confirmed"):
        return False
    if _docs_gate_items(snapshot):
        return False
    if snapshot.get("qa_status") in {"qa_passed", "has_issues"}:
        return True
    return False


def _docs_gate_items(snapshot: dict[str, Any]) -> list[dict[str, str]]:
    """Doc-gate-blocking items from missing_items: everything except QA kind.

    Filters by ``kind`` for structured entries; falls back to the legacy
    "QA " text-prefix rule for any string entries still in stale snapshots.
    """
    return [item for item in snapshot.get("missing_items", []) if missing_item_kind(item) != "qa"]


def md_title(title: str, level: int = 1) -> str:
    hashes = "#" * max(1, min(6, level))
    return f"{hashes} {title}\n\n"


def code_block(content: str, lang: str = "shell") -> str:
    if not content:
        return "\n"
    return f"```{lang}\n{content}\n```\n\n"


def _md_cell(value: Any) -> str:
    s = "" if value is None else str(value)
    s = s.replace("\\", "\\\\").replace("|", "\\|")
    return s.replace("\r\n", " ").replace("\n", "<br>").replace("\r", " ")


def release_rows(conn: sqlite3.Connection, release: dict[str, Any], *, final: bool = False) -> list[tuple[dict[str, Any], dict[str, Any]]]:
    """Rows for the release note.

    Preview and final RST both include only apps that currently qualify for
    release. Unfinished apps stay visible in missing_items, not generated RST.
    """
    apps = {app["id"]: app for app in list_apps(conn)}
    rows = []
    for app_id, snapshot in release["snapshots"].items():
        app = apps.get(app_id)
        if not app or snapshot.get("release_decision") != "release":
            continue
        if not _qualifies_for_final(snapshot):
            continue
        rows.append((app_view(app, snapshot), snapshot))
    return sorted(rows, key=lambda item: item[0]["name"].lower())


def guide_rows(
    conn: sqlite3.Connection,
    release: dict[str, Any],
    doc_target: str,
) -> tuple[list[tuple[dict[str, Any], dict[str, Any]]], list[tuple[dict[str, Any], dict[str, Any]]]]:
    """Return ``(active_rows, stopped_rows)`` for a manual/ai4sci guide."""
    apps = {app["id"]: app for app in list_apps(conn)}
    active: list[tuple[dict[str, Any], dict[str, Any]]] = []
    stopped: list[tuple[dict[str, Any], dict[str, Any]]] = []

    for app_id, snapshot in release["snapshots"].items():
        app = apps.get(app_id)
        if not app:
            continue
        if normalize_doc_target(snapshot.get("doc_target")) != doc_target:
            continue

        decision = snapshot.get("release_decision", "release")
        qualifies = _qualifies_for_final(snapshot)

        if decision == "stopped":
            prev = last_published_snapshot(conn, app_id, release["id"])
            if prev:
                stopped.append((app_view(app, prev), prev))
            continue

        if decision != "release":
            continue

        prev = last_published_snapshot(conn, app_id, release["id"])

        if qualifies:
            active.append((app_view(app, snapshot), snapshot))
        elif prev:
            active.append((app_view(app, prev), prev))
        # else: new app that didn't qualify — omit

    active.sort(key=lambda item: item[0]["name"].lower())
    stopped.sort(key=lambda item: item[0]["name"].lower())
    return active, stopped


def render_release_note(release: dict[str, Any], rows: list[tuple[dict[str, Any], dict[str, Any]]]) -> str:
    out = md_title(f"MACA HPC 发布列表 - {release['name']}")
    headers = ["名称", "类型", "描述", "对应官方版本", "X86支持芯片系列", "ARM支持芯片类型"]
    out += "| " + " | ".join(headers) + " |\n"
    out += "| " + " | ".join("---" for _ in headers) + " |\n"
    for app, snapshot in rows:
        cells = [
            app["name"],
            app.get("type") or "",
            app.get("description") or "",
            snapshot.get("version") or "",
            snapshot.get("x86_chips") or "",
            snapshot.get("arm_chips") or "",
        ]
        out += "| " + " | ".join(_md_cell(c) for c in cells) + " |\n"
    out += "\n"
    return out


def _merged_limitations(snapshot: dict[str, Any]) -> str:
    """Merge owner-written limitations with QA's issue note if any."""
    text = (snapshot.get("doc", {}) or {}).get("limitations", "") or ""
    if snapshot.get("qa_status") == "has_issues" and snapshot.get("qa_issue_note"):
        prefix = f"QA 备注：{snapshot['qa_issue_note']}"
        text = f"{text}\n\n{prefix}".strip() if text else prefix
    return text


def _chip_support_text(snapshot: dict[str, Any]) -> str:
    parts = []
    if snapshot.get("x86_chips"):
        parts.append(f"X86: {snapshot['x86_chips']}")
    if snapshot.get("arm_chips"):
        parts.append(f"ARM: {snapshot['arm_chips']}")
    if snapshot.get("hpcc_chip"):
        parts.append(f"HPCC: {snapshot['hpcc_chip']}")
    return "; ".join(parts)


def _not_releasable_reason(snapshot: dict[str, Any]) -> str:
    decision = normalize_release_decision(snapshot.get("release_decision"))
    if _qualifies_for_final(snapshot):
        return ""
    if decision != "release":
        return f"Release决策为 {decision}"
    missing = snapshot.get("missing_items", [])
    return "；".join(missing_item_text(m) for m in missing) if missing else "未满足发布条件"


def render_manager_review_csv(
    conn: sqlite3.Connection,
    release: dict[str, Any],
    fields: list[str] | None = None,
) -> str:
    """Return manager-review CSV for all apps in the release snapshots."""
    field_labels = dict(MANAGER_REVIEW_FIELDS)
    selected = fields or DEFAULT_MANAGER_REVIEW_FIELDS
    if not selected:
        raise ValueError("至少选择一个输出字段")
    invalid = [field for field in selected if field not in field_labels]
    if invalid:
        raise ValueError(f"未知 Manager Review 字段: {', '.join(invalid)}")

    apps = {app["id"]: app for app in list_apps(conn)}
    out = io.StringIO()
    writer = csv.writer(out)
    writer.writerow([field_labels[field] for field in selected])
    rows = []
    for app_id, snapshot in release["snapshots"].items():
        app = apps.get(app_id)
        if not app:
            continue
        rows.append((app_view(app, snapshot), snapshot))
    rows.sort(key=lambda item: item[0].get("name", "").lower())
    for app, snapshot in rows:
        values = {
            "app_name": app.get("name", ""),
            "official_name": app.get("official_name", ""),
            "doc_target": "AI4Sci" if normalize_doc_target(app.get("doc_target")) == "ai4sci" else "HPC",
            "app_type": app.get("type", ""),
            "version": snapshot.get("version", ""),
            "owners": ",".join(app.get("owners", [])),
            "chip_support": _chip_support_text(snapshot),
            "x86_chips": snapshot.get("x86_chips", ""),
            "arm_chips": snapshot.get("arm_chips", ""),
            "release_decision": normalize_release_decision(snapshot.get("release_decision")),
            "qa_status": snapshot.get("qa_status", "not_checked"),
            "owner_confirmed": "是" if snapshot.get("owner_confirmed") else "否",
            "releasable": "是" if _qualifies_for_final(snapshot) else "否",
            "not_releasable_reason": _not_releasable_reason(snapshot),
            "known_limitations": _merged_limitations(snapshot),
            "gerrit_url": app.get("git_url", ""),
            "git_branch": app.get("git_branch", ""),
        }
        writer.writerow([values[field] for field in selected])
    return out.getvalue()


def _render_guide_entries(rows: list[tuple[dict[str, Any], dict[str, Any]]], *, stopped: bool = False) -> str:
    out = ""
    for app, snapshot in rows:
        doc = snapshot.get("doc", {})
        heading = f"{app['name']}（已停止支持）" if stopped else app["name"]
        out += md_title(heading, 2)
        out += f"{doc.get('intro') or app.get('description') or ''}\n\n"
        out += f"版本：{snapshot.get('version') or ''}\n\n"
        if app.get("official_url"):
            out += f"官方网址：{app['official_url']}\n\n"
        out += "**镜像使用方法：**\n\n" + code_block(doc.get("image_usage", ""))
        out += "**二进制包使用方法：**\n\n" + code_block(doc.get("binary_usage", ""))
        out += "**环境搭建：**\n\n" + code_block(doc.get("env_setup", ""))
        out += "**测试方法：**\n\n"
        for test_doc in snapshot.get("test_docs", []):
            if test_doc.get("obsolete"):
                continue
            out += f"- {test_doc['path']}\n"
            out += f"  - 测试数据集：{test_doc.get('dataset', '')}\n"
            out += f"  - 测试内容：{test_doc.get('content', '')}\n"
            out += f"  - 结果查看：{test_doc.get('result_view', '')}\n"
            out += f"  - 通过标准：{test_doc.get('pass_criteria', '')}\n\n"
            if test_doc.get("command"):
                out += code_block(test_doc["command"])
        limits = _merged_limitations(snapshot)
        if limits:
            out += f"**已知限制：**\n\n{limits}\n\n"
    return out


def render_guide(
    title: str,
    rows: list[tuple[dict[str, Any], dict[str, Any]]],
    stopped_rows: list[tuple[dict[str, Any], dict[str, Any]]] | None = None,
) -> str:
    out = md_title(title)
    out += _render_guide_entries(rows)
    if stopped_rows:
        out += _render_guide_entries(stopped_rows, stopped=True)
    return out


def generate_artifacts(conn: sqlite3.Connection, release_id: str, *, final: bool = False, from_lock: bool = False) -> dict[str, str]:
    release = get_release(conn, release_id)
    if final and not from_lock:
        raise RuntimeError("Final artifacts 只能通过 final_lock_release 生成")
    if from_lock and not final:
        raise RuntimeError("Lock generation must create final artifacts")
    if from_lock and not release.get("released_locked"):
        raise RuntimeError("Final artifacts require a locked release")
    if release.get("released_locked") and not from_lock:
        raise RuntimeError("Release 已最终锁定，artifacts 不可重新生成")
    if final:
        existing = conn.execute("SELECT 1 FROM artifacts WHERE release_id = ? AND final = 1 LIMIT 1", (release_id,)).fetchone()
        if existing:
            raise RuntimeError("Final artifacts already exist and are immutable")
    with transaction(conn):
        if not final:
            refresh_missing_items(conn, release_id)
            release = get_release(conn, release_id)
        rows = release_rows(conn, release, final=final)
        if final:
            manual_active, manual_stopped = guide_rows(conn, release, "manual")
            ai4sci_active, ai4sci_stopped = guide_rows(conn, release, "ai4sci")
        else:
            manual_active = [(a, s) for a, s in rows if normalize_doc_target(a.get("doc_target")) == "manual"]
            manual_stopped = []
            ai4sci_active = [(a, s) for a, s in rows if normalize_doc_target(a.get("doc_target")) == "ai4sci"]
            ai4sci_stopped = []
        artifacts = {
            "release_note": render_release_note(release, rows),
            "manual": render_guide("HPC Manual App 章节", manual_active, manual_stopped),
            "ai4sci": render_guide("AI4Sci User Guide App 章节", ai4sci_active, ai4sci_stopped),
            "data": dumps_json({"release": release, "apps": list_apps(conn), "generated_at": now(), "final": final}),
        }
        names = {
            "release_note": "release_note.md",
            "manual": "hpc_manual_apps.md",
            "ai4sci": "ai4sci_user_guide_apps.md",
            "data": "release_data.json",
        }
        for kind, content in artifacts.items():
            conn.execute(
                """
                INSERT INTO artifacts(release_id, kind, name, content, final, generated_at)
                VALUES (?, ?, ?, ?, ?, ?)
                ON CONFLICT(release_id, kind) DO UPDATE SET
                  name=excluded.name,
                  content=excluded.content,
                  final=excluded.final,
                  generated_at=excluded.generated_at
                """,
                (release_id, kind, names[kind], content, int(final), now()),
            )
        audit(
            conn,
            "生成最终 artifacts" if final else "生成预览 artifacts",
            release_id=release_id,
            event="generate_artifacts",
        )
    return artifacts


def generate_manager_review_csv(
    conn: sqlite3.Connection,
    release_id: str,
    fields: list[str] | None = None,
    *,
    user: str = "rm",
    role: str = "RM",
) -> str:
    release = get_release(conn, release_id)
    if release.get("released_locked"):
        raise RuntimeError("Release 已最终锁定，Manager Review CSV 不可重新生成")
    with transaction(conn):
        refresh_missing_items(conn, release_id)
        release = get_release(conn, release_id)
        content = render_manager_review_csv(conn, release, fields)
        conn.execute(
            """
            INSERT INTO artifacts(release_id, kind, name, content, final, generated_at)
            VALUES (?, ?, ?, ?, ?, ?)
            ON CONFLICT(release_id, kind) DO UPDATE SET
              name=excluded.name,
              content=excluded.content,
              final=excluded.final,
              generated_at=excluded.generated_at
            """,
            (release_id, "manager_review", "manager_review.csv", content, 0, now()),
        )
        audit(
            conn,
            "生成 Manager Review CSV",
            user=user,
            role=role,
            release_id=release_id,
            event="generate_manager_review",
        )
    return content


def gerrit_push_plan(conn: sqlite3.Connection, release_id: str) -> dict[str, Any]:
    release = get_release(conn, release_id)
    if not release.get("released_locked"):
        raise RuntimeError("Gerrit push 要求 release 已最终锁定")
    docs_remote = os.environ.get("HPC_DOCS_GERRIT_REMOTE", "")
    data_remote = os.environ.get("HPC_RELEASE_DATA_GERRIT_REMOTE", "")
    if not docs_remote or not data_remote:
        return {
            "ready": False,
            "reason": "Missing HPC_DOCS_GERRIT_REMOTE or HPC_RELEASE_DATA_GERRIT_REMOTE",
            "required_env": ["HPC_DOCS_GERRIT_REMOTE", "HPC_RELEASE_DATA_GERRIT_REMOTE"],
        }
    branch = f"release-{release['name']}"
    return {
        "ready": True,
        "docs_remote": docs_remote,
        "data_remote": data_remote,
        "branch": branch,
        "commands": [
            f"git clone {docs_remote} docs-worktree",
            f"git -C docs-worktree checkout -b {branch}",
            "copy generated Markdown artifacts into docs-worktree",
            f"git -C docs-worktree push origin HEAD:refs/for/{branch}",
            f"git clone {data_remote} release-data-worktree",
            f"git -C release-data-worktree checkout -b {branch}",
            "copy release_data.json into release-data-worktree",
            f"git -C release-data-worktree push origin HEAD:refs/for/{branch}",
        ],
    }


# ---------------------------------------------------------------------------
# CICD workbench
# ---------------------------------------------------------------------------

CICD_BUILD_PRODUCTS = ["maca", "hpcc", "pkg"]
CICD_STATUSES = {"Running", "Stopped", "Abandoned"}
CICD_APPROVER_ROLES = {"RM", "Admin"}
CICD_CREATE_ROLES = {"Owner", "RM", "Admin"}


def _next_cicd_id(conn: sqlite3.Connection) -> str:
    row = conn.execute("SELECT id FROM cicd_tasks ORDER BY id DESC LIMIT 1").fetchone()
    if not row:
        return "CICD-0001"
    last = row["id"]  # e.g. CICD-0042
    try:
        num = int(last.split("-", 1)[1]) + 1
    except (IndexError, ValueError):
        num = 1
    return f"CICD-{num:04d}"


def _cicd_task_row(row) -> dict:
    d = dict(row)
    try:
        d["build_product"] = json.loads(d.get("build_product") or "[]")
    except Exception:
        d["build_product"] = []
    try:
        d["community_artifact"] = json.loads(d.get("community_artifact") or "[]")
    except Exception:
        d["community_artifact"] = []
    return d


def list_cicd_tasks(
    conn: sqlite3.Connection,
    *,
    status_filter: str | None = None,   # "Running" | "Stopped" | "Abandoned" | None=all
) -> list[dict]:
    if status_filter and status_filter in CICD_STATUSES:
        rows = conn.execute(
            "SELECT * FROM cicd_tasks WHERE status = ? ORDER BY id",
            (status_filter,),
        ).fetchall()
    else:
        rows = conn.execute("SELECT * FROM cicd_tasks ORDER BY id").fetchall()
    tasks = [_cicd_task_row(r) for r in rows]
    # attach pending-request flag for each task
    pending_task_ids = {
        r["task_id"]
        for r in conn.execute(
            "SELECT DISTINCT task_id FROM cicd_task_requests WHERE status = 'pending' AND task_id IS NOT NULL"
        ).fetchall()
    }
    delivery_pending_ids = {
        r["task_id"]
        for r in conn.execute(
            "SELECT DISTINCT task_id FROM cicd_task_requests "
            "WHERE delivery_status IN ('pending', 'returned') AND task_id IS NOT NULL"
        ).fetchall()
    }
    for t in tasks:
        t["has_pending"] = t["id"] in pending_task_ids
        t["has_pending_delivery"] = t["id"] in delivery_pending_ids
    # attach owner display_name
    _attach_owner_display(conn, tasks)
    return tasks


def _attach_owner_display(conn: sqlite3.Connection, tasks: list[dict]) -> None:
    if not tasks:
        return
    usernames = list({t["owner_username"] for t in tasks})
    rows = conn.execute(
        "SELECT username, display_name FROM users WHERE username IN ({})".format(
            ",".join("?" * len(usernames))
        ),
        usernames,
    ).fetchall()
    display_map = {r["username"]: r["display_name"] for r in rows}
    for t in tasks:
        u = t["owner_username"]
        dn = display_map.get(u, "")
        t["owner_display"] = f"{dn}" if dn else u


def get_cicd_task(conn: sqlite3.Connection, task_id: str) -> dict | None:
    row = conn.execute("SELECT * FROM cicd_tasks WHERE id = ?", (task_id,)).fetchone()
    if not row:
        return None
    task = _cicd_task_row(row)
    _attach_owner_display(conn, [task])
    return task


def submit_cicd_request(
    conn: sqlite3.Connection,
    *,
    task_id: str | None,
    request_type: str,           # "create" | "modify"
    payload: dict,               # {field: {old, new}} for modify; full fields for create
    submitter: str,
    submitter_role: str,
    submitter_display: str = "",
) -> dict:
    """Submit a CICD task create/modify request.

    For RM/Admin submitters: auto-approves immediately and applies the change.
    For Owner submitters: enters the pending queue.
    Returns the created request as a dict.
    """
    if submitter_role not in CICD_CREATE_ROLES:
        raise PermissionError("只有 Owner、RM、Admin 可以提交 CICD 任务申请")
    is_auto = submitter_role in CICD_APPROVER_ROLES
    ts = now()
    with transaction(conn):
        conn.execute(
            """
            INSERT INTO cicd_task_requests
              (task_id, request_type, payload, submitter, submitter_display,
               submitted_at, status, reviewer, reviewed_at, review_note, is_self_approved)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                task_id,
                request_type,
                json.dumps(payload, ensure_ascii=False),
                submitter,
                submitter_display,
                ts,
                "approved" if is_auto else "pending",
                submitter if is_auto else "",
                ts if is_auto else "",
                "",
                1 if is_auto else 0,
            ),
        )
        req_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        if is_auto:
            _apply_cicd_request(conn, req_id, payload, task_id, request_type, ts)
        row = conn.execute("SELECT * FROM cicd_task_requests WHERE id = ?", (req_id,)).fetchone()
    return dict(row)


def _apply_cicd_request(
    conn: sqlite3.Connection,
    req_id: int,
    payload: dict,
    task_id: str | None,
    request_type: str,
    ts: str,
) -> str:
    """Actually create or update cicd_tasks after approval. Returns the task_id."""
    if request_type == "create":
        new_id = _next_cicd_id(conn)
        build_product = payload.get("build_product", [])
        community_artifact = payload.get("community_artifact", [])
        conn.execute(
            """
            INSERT INTO cicd_tasks
              (id, app_name, app_version, repo_type, repo_name, branch,
               build_product, community_artifact, build_image, test_timeout, supports_maca_hpcc,
               owner_username, status, notes, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                new_id,
                payload.get("app_name", ""),
                payload.get("app_version", ""),
                payload.get("repo_type", "git"),
                payload.get("repo_name", ""),
                payload.get("branch", ""),
                json.dumps(build_product, ensure_ascii=False),
                json.dumps(community_artifact, ensure_ascii=False),
                payload.get("build_image", ""),
                int(payload.get("test_timeout", 40)),
                payload.get("supports_maca_hpcc", "No"),
                payload.get("owner_username", ""),
                payload.get("status", "Running"),
                payload.get("notes", ""),
                ts,
                ts,
            ),
        )
        # back-fill task_id on the request row
        conn.execute(
            "UPDATE cicd_task_requests SET task_id = ? WHERE id = ?",
            (new_id, req_id),
        )
        return new_id
    else:
        # modify: apply diff
        if not task_id:
            raise RuntimeError("修改请求缺少 task_id")
        for field, change in payload.items():
            new_val = change.get("new")
            if field == "build_product":
                new_val = json.dumps(new_val or [], ensure_ascii=False)
            elif field == "community_artifact":
                new_val = json.dumps(new_val or [], ensure_ascii=False)
            elif field == "test_timeout":
                new_val = int(new_val or 40)
            conn.execute(
                f"UPDATE cicd_tasks SET {field} = ?, updated_at = ? WHERE id = ?",
                (new_val, ts, task_id),
            )
        return task_id


def approve_cicd_request(
    conn: sqlite3.Connection,
    req_id: int,
    *,
    reviewer: str,
    reviewer_role: str,
    review_note: str = "",
    approval_mode: str = "immediate",   # "immediate" | "dispatch_spd"
    jira_id: str = "",
    jira_auto_created: int = 0,
) -> dict:
    """Approve a pending CICD request.

    approval_mode='immediate': apply change right away (立即生效).
    approval_mode='dispatch_spd': defer apply until SPD delivers (批准并下发).
    """
    if reviewer_role not in CICD_APPROVER_ROLES:
        raise PermissionError("只有 RM/Admin 可以审批 CICD 任务申请")
    row = conn.execute("SELECT * FROM cicd_task_requests WHERE id = ?", (req_id,)).fetchone()
    if not row:
        raise RuntimeError("申请不存在")
    req = dict(row)
    if req["status"] != "pending":
        raise RuntimeError(f"申请状态为 {req['status']}，无法审批")
    ts = now()
    payload = json.loads(req["payload"] or "{}")
    with transaction(conn):
        if approval_mode == "dispatch_spd":
            conn.execute(
                """UPDATE cicd_task_requests
                   SET status='approved', reviewer=?, reviewed_at=?, review_note=?,
                       approval_mode='dispatch_spd', delivery_status='pending',
                       jira_id=?, jira_auto_created=?
                   WHERE id=?""",
                (reviewer, ts, review_note, jira_id, jira_auto_created, req_id),
            )
            # _apply_cicd_request intentionally deferred until SPD delivers
        else:
            conn.execute(
                """UPDATE cicd_task_requests
                   SET status='approved', reviewer=?, reviewed_at=?, review_note=?,
                       approval_mode='immediate'
                   WHERE id=?""",
                (reviewer, ts, review_note, req_id),
            )
            _apply_cicd_request(conn, req_id, payload, req["task_id"], req["request_type"], ts)
        row = conn.execute("SELECT * FROM cicd_task_requests WHERE id = ?", (req_id,)).fetchone()
    return dict(row)


def reject_cicd_request(
    conn: sqlite3.Connection,
    req_id: int,
    *,
    reviewer: str,
    reviewer_role: str,
    review_note: str,
) -> dict:
    if reviewer_role not in CICD_APPROVER_ROLES:
        raise PermissionError("只有 RM/Admin 可以拒绝 CICD 任务申请")
    if not review_note or not review_note.strip():
        raise ValueError("拒绝必须填写理由")
    row = conn.execute("SELECT * FROM cicd_task_requests WHERE id = ?", (req_id,)).fetchone()
    if not row:
        raise RuntimeError("申请不存在")
    if dict(row)["status"] != "pending":
        raise RuntimeError(f"申请状态为 {dict(row)['status']}，无法拒绝")
    ts = now()
    with transaction(conn):
        conn.execute(
            "UPDATE cicd_task_requests SET status='rejected', reviewer=?, reviewed_at=?, review_note=? WHERE id=?",
            (reviewer, ts, review_note.strip(), req_id),
        )
        row = conn.execute("SELECT * FROM cicd_task_requests WHERE id = ?", (req_id,)).fetchone()
    return dict(row)


def cancel_cicd_request(
    conn: sqlite3.Connection,
    req_id: int,
    *,
    username: str,
    role: str,
) -> dict:
    row = conn.execute("SELECT * FROM cicd_task_requests WHERE id = ?", (req_id,)).fetchone()
    if not row:
        raise RuntimeError("申请不存在")
    req = dict(row)
    if req["status"] != "pending":
        raise RuntimeError(f"申请状态为 {req['status']}，只有 pending 状态可以取消")
    # Only submitter or RM/Admin can cancel
    if req["submitter"] != username and role not in CICD_APPROVER_ROLES:
        raise PermissionError("只有提交人或 RM/Admin 可以取消申请")
    with transaction(conn):
        conn.execute(
            "UPDATE cicd_task_requests SET status='cancelled', reviewed_at=? WHERE id=?",
            (now(), req_id),
        )
        row = conn.execute("SELECT * FROM cicd_task_requests WHERE id = ?", (req_id,)).fetchone()
    return dict(row)


def transfer_cicd_owner(
    conn: sqlite3.Connection,
    task_id: str,
    new_owner: str,
    *,
    actor: str,
    actor_role: str,
) -> dict:
    """RM/Admin can transfer task ownership directly without approval flow."""
    if actor_role not in CICD_APPROVER_ROLES:
        raise PermissionError("只有 RM/Admin 可以直接修改负责人")
    task = get_cicd_task(conn, task_id)
    if not task:
        raise RuntimeError(f"CICD 任务 {task_id} 不存在")
    old_owner = task["owner_username"]
    ts = now()
    with transaction(conn):
        conn.execute(
            "UPDATE cicd_tasks SET owner_username=?, updated_at=? WHERE id=?",
            (new_owner, ts, task_id),
        )
        # Record in request history as a special "owner_transfer" entry
        payload = json.dumps(
            {"owner_username": {"old": old_owner, "new": new_owner}},
            ensure_ascii=False,
        )
        conn.execute(
            """
            INSERT INTO cicd_task_requests
              (task_id, request_type, payload, submitter, submitter_display,
               submitted_at, status, reviewer, reviewed_at, review_note, is_self_approved)
            VALUES (?, 'owner_transfer', ?, ?, ?, ?, 'approved', ?, ?, '负责人直接变更', 1)
            """,
            (task_id, payload, actor, actor, ts, actor, ts),
        )
    return get_cicd_task(conn, task_id)


def delete_cicd_task(
    conn: sqlite3.Connection,
    task_id: str,
    *,
    actor: str,
    actor_role: str,
) -> None:
    if actor_role not in CICD_APPROVER_ROLES:
        raise PermissionError("只有 RM/Admin 可以删除 CICD 任务")
    task = get_cicd_task(conn, task_id)
    if not task:
        raise RuntimeError(f"CICD 任务 {task_id} 不存在")
    if task["status"] != "Abandoned":
        raise RuntimeError("只有 Abandoned 状态的任务可以删除")
    with transaction(conn):
        conn.execute("DELETE FROM cicd_task_requests WHERE task_id = ?", (task_id,))
        conn.execute("DELETE FROM cicd_tasks WHERE id = ?", (task_id,))


def list_cicd_requests(
    conn: sqlite3.Connection,
    *,
    username: str | None = None,   # filter by submitter (None = all)
    role: str = "Owner",
    task_id: str | None = None,
    status_filter: str | None = None,   # pending/approved/rejected/cancelled/None=all
    since_days: int | None = None,       # filter by time window
    exclude_cancelled: bool = False,
) -> list[dict]:
    """Return cicd_task_requests with flexible filters.

    Visibility rules:
    - RM/Admin: can see all records (except if username filter applied)
    - Owner: can only see their own submissions; cancelled records hidden from
      non-RM/Admin when exclude_cancelled=True
    """
    clauses: list[str] = []
    params: list = []

    if task_id:
        clauses.append("task_id = ?")
        params.append(task_id)

    if status_filter:
        clauses.append("status = ?")
        params.append(status_filter)

    if exclude_cancelled and role not in CICD_APPROVER_ROLES:
        clauses.append("(status != 'cancelled' OR submitter = ?)")
        params.append(username or "")

    if username and role not in CICD_APPROVER_ROLES:
        # Non-RM/Admin can only see their own
        clauses.append("submitter = ?")
        params.append(username)
    elif username:
        # RM/Admin with explicit username filter (e.g. "only mine" toggle)
        clauses.append("submitter = ?")
        params.append(username)

    if since_days:
        cutoff = conn.execute(
            "SELECT datetime('now', ?)", (f"-{since_days} days",)
        ).fetchone()[0]
        clauses.append("submitted_at >= ?")
        params.append(cutoff)

    where = ("WHERE " + " AND ".join(clauses)) if clauses else ""
    rows = conn.execute(
        f"SELECT * FROM cicd_task_requests {where} ORDER BY submitted_at DESC",
        params,
    ).fetchall()
    result = []
    for r in rows:
        d = dict(r)
        try:
            d["payload"] = json.loads(d.get("payload") or "{}")
        except Exception:
            d["payload"] = {}
        result.append(d)
    return result


def get_cicd_task_history(conn: sqlite3.Connection, task_id: str) -> list[dict]:
    """Return all approved/auto-approved requests for a specific task (chronological)."""
    rows = conn.execute(
        """
        SELECT * FROM cicd_task_requests
        WHERE task_id = ? AND status = 'approved'
        ORDER BY reviewed_at ASC
        """,
        (task_id,),
    ).fetchall()
    result = []
    for r in rows:
        d = dict(r)
        try:
            d["payload"] = json.loads(d.get("payload") or "{}")
        except Exception:
            d["payload"] = {}
        result.append(d)
    return result


def get_cicd_notifications(
    conn: sqlite3.Connection,
    username: str,
    role: str,
) -> dict:
    """Return notification counts for the nav badge."""
    row = conn.execute(
        "SELECT last_visited_at FROM cicd_notifications WHERE username = ?",
        (username,),
    ).fetchone()
    last_visited = row["last_visited_at"] if row else ""

    if role == "SPD":
        # SPD: count requests awaiting delivery
        count = conn.execute(
            "SELECT COUNT(*) FROM cicd_task_requests WHERE delivery_status = 'pending'",
        ).fetchone()[0]
    elif role in CICD_APPROVER_ROLES:
        # RM/Admin: pending approval + returned by SPD (both need action)
        pending = conn.execute(
            "SELECT COUNT(*) FROM cicd_task_requests WHERE status = 'pending'",
        ).fetchone()[0]
        returned = conn.execute(
            "SELECT COUNT(*) FROM cicd_task_requests WHERE delivery_status = 'returned'",
        ).fetchone()[0]
        count = pending + returned
    else:
        # Owner: new approved/rejected since last visit
        if last_visited:
            count = conn.execute(
                """
                SELECT COUNT(*) FROM cicd_task_requests
                WHERE submitter = ? AND status IN ('approved', 'rejected')
                AND reviewed_at > ?
                """,
                (username, last_visited),
            ).fetchone()[0]
        else:
            count = 0

    return {"count": count, "last_visited_at": last_visited}


def mark_cicd_visited(conn: sqlite3.Connection, username: str) -> None:
    """Update the last_visited timestamp for notification badge."""
    ts = now()
    with transaction(conn):
        conn.execute(
            """
            INSERT INTO cicd_notifications(username, last_visited_at)
            VALUES (?, ?)
            ON CONFLICT(username) DO UPDATE SET last_visited_at=excluded.last_visited_at
            """,
            (username, ts),
        )


# ---------------------------------------------------------------------------
# Delivery workflow functions (SPD path)
# ---------------------------------------------------------------------------

def deliver_cicd_request(
    conn: sqlite3.Connection,
    req_id: int,
    *,
    deliverer: str,
    deliverer_role: str,
) -> dict:
    """SPD (or RM/Admin) marks a dispatched request as delivered.
    Applies the change to cicd_tasks at this point.
    """
    if deliverer_role not in {"SPD", "RM", "Admin"}:
        raise PermissionError("只有 SPD、RM、Admin 可以标记已交付")
    row = conn.execute("SELECT * FROM cicd_task_requests WHERE id = ?", (req_id,)).fetchone()
    if not row:
        raise RuntimeError("申请不存在")
    req = dict(row)
    if req.get("delivery_status") not in ("pending", "returned"):
        raise RuntimeError(f"该申请的交付状态为 '{req.get('delivery_status')}'，无法标记已交付")
    ts = now()
    payload = json.loads(req["payload"] or "{}")
    with transaction(conn):
        _apply_cicd_request(conn, req_id, payload, req["task_id"], req["request_type"], ts)
        conn.execute(
            """UPDATE cicd_task_requests
               SET delivery_status='delivered', delivered_by=?, delivered_at=?
               WHERE id=?""",
            (deliverer, ts, req_id),
        )
        row = conn.execute("SELECT * FROM cicd_task_requests WHERE id = ?", (req_id,)).fetchone()
    return dict(row)


def return_cicd_request(
    conn: sqlite3.Connection,
    req_id: int,
    *,
    returner: str,
    returner_role: str,
    reason: str,
) -> dict:
    """SPD returns a delivery back to RM with a reason."""
    if returner_role != "SPD":
        raise PermissionError("只有 SPD 可以退回交付申请")
    if not reason or not reason.strip():
        raise ValueError("退回必须填写原因")
    row = conn.execute("SELECT * FROM cicd_task_requests WHERE id = ?", (req_id,)).fetchone()
    if not row:
        raise RuntimeError("申请不存在")
    req = dict(row)
    if req.get("delivery_status") != "pending":
        raise RuntimeError(f"该申请的交付状态为 '{req.get('delivery_status')}'，无法退回")
    ts = now()
    with transaction(conn):
        conn.execute(
            """UPDATE cicd_task_requests
               SET delivery_status='returned', returned_reason=?, returned_at=?
               WHERE id=?""",
            (reason.strip(), ts, req_id),
        )
        row = conn.execute("SELECT * FROM cicd_task_requests WHERE id = ?", (req_id,)).fetchone()
    return dict(row)


def re_dispatch_cicd_request(
    conn: sqlite3.Connection,
    req_id: int,
    *,
    actor: str,
    actor_role: str,
) -> dict:
    """RM/Admin re-dispatches a returned delivery back to SPD."""
    if actor_role not in CICD_APPROVER_ROLES:
        raise PermissionError("只有 RM/Admin 可以重新下发")
    row = conn.execute("SELECT * FROM cicd_task_requests WHERE id = ?", (req_id,)).fetchone()
    if not row:
        raise RuntimeError("申请不存在")
    req = dict(row)
    if req.get("delivery_status") != "returned":
        raise RuntimeError(f"该申请的交付状态为 '{req.get('delivery_status')}'，只有 returned 状态可以重新下发")
    with transaction(conn):
        conn.execute(
            "UPDATE cicd_task_requests SET delivery_status='pending' WHERE id=?",
            (req_id,),
        )
        row = conn.execute("SELECT * FROM cicd_task_requests WHERE id = ?", (req_id,)).fetchone()
    return dict(row)


def apply_returned_cicd_request(
    conn: sqlite3.Connection,
    req_id: int,
    *,
    actor: str,
    actor_role: str,
) -> dict:
    """RM/Admin decides to apply a returned (or pending) request immediately, bypassing SPD."""
    if actor_role not in CICD_APPROVER_ROLES:
        raise PermissionError("只有 RM/Admin 可以直接生效")
    row = conn.execute("SELECT * FROM cicd_task_requests WHERE id = ?", (req_id,)).fetchone()
    if not row:
        raise RuntimeError("申请不存在")
    req = dict(row)
    if req.get("delivery_status") not in ("pending", "returned"):
        raise RuntimeError(f"该申请的交付状态为 '{req.get('delivery_status')}'，无法直接生效")
    ts = now()
    payload = json.loads(req["payload"] or "{}")
    with transaction(conn):
        _apply_cicd_request(conn, req_id, payload, req["task_id"], req["request_type"], ts)
        conn.execute(
            """UPDATE cicd_task_requests
               SET delivery_status='delivered', delivered_by=?, delivered_at=?,
                   approval_mode='immediate'
               WHERE id=?""",
            (actor, ts, req_id),
        )
        row = conn.execute("SELECT * FROM cicd_task_requests WHERE id = ?", (req_id,)).fetchone()
    return dict(row)


def list_cicd_deliveries(
    conn: sqlite3.Connection,
    *,
    status_filter: str | None = None,
    role: str = "SPD",
    submitter: str | None = None,
) -> list[dict]:
    """List requests that went through the dispatch_spd delivery workflow.

    status_filter:
      'pending'            — awaiting SPD delivery
      'returned'           — SPD returned, RM needs to act
      'pending_or_returned'— both (RM/Owner view)
      'delivered'          — completed deliveries (history)
      None                 — all dispatch_spd requests
    submitter:
      When set, restricts results to a single submitter (used for Owner role).
    """
    clauses = ["approval_mode = 'dispatch_spd'"]
    params: list = []

    if status_filter == "pending_or_returned":
        clauses.append("delivery_status IN ('pending', 'returned')")
    elif status_filter:
        clauses.append("delivery_status = ?")
        params.append(status_filter)

    if submitter:
        clauses.append("submitter = ?")
        params.append(submitter)

    where = "WHERE " + " AND ".join(clauses)
    rows = conn.execute(
        f"SELECT * FROM cicd_task_requests {where} ORDER BY reviewed_at DESC",
        params,
    ).fetchall()

    result = []
    for r in rows:
        d = dict(r)
        try:
            d["payload"] = json.loads(d.get("payload") or "{}")
        except Exception:
            d["payload"] = {}
        result.append(d)
    _attach_delivery_task_info(conn, result)
    return result


def _attach_delivery_task_info(conn: sqlite3.Connection, deliveries: list[dict]) -> None:
    """Attach current task app_name/app_version/status to each delivery record."""
    task_ids = [d["task_id"] for d in deliveries if d.get("task_id")]
    task_map: dict = {}
    if task_ids:
        rows = conn.execute(
            "SELECT id, app_name, app_version, status FROM cicd_tasks WHERE id IN ({})".format(
                ",".join("?" * len(task_ids))
            ),
            task_ids,
        ).fetchall()
        task_map = {r["id"]: dict(r) for r in rows}
    for d in deliveries:
        t = task_map.get(d.get("task_id") or "", {})
        if not t and d.get("request_type") == "create":
            # task not yet created; get display info from the create payload
            p = d.get("payload") or {}
            d["task_app_name"] = p.get("app_name", "")
            d["task_app_version"] = p.get("app_version", "")
            d["task_status"] = p.get("status", "Running")
        else:
            d["task_app_name"] = t.get("app_name", "")
            d["task_app_version"] = t.get("app_version", "")
            d["task_status"] = t.get("status", "")
